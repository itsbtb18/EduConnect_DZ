"""
Attendance views — teachers mark, admins review, parents/students view.
"""

import csv
import io
from datetime import timedelta

from django.db.models import Count, Q
from django.http import HttpResponse
from django.shortcuts import get_object_or_404
from django.utils import timezone
from drf_spectacular.utils import OpenApiResponse, extend_schema, inline_serializer
from rest_framework import permissions, serializers, status
from rest_framework.response import Response
from rest_framework.views import APIView

from apps.academics.models import Class, StudentProfile, TeacherAssignment
from apps.accounts.models import User
from core.permissions import IsParent, IsSchoolAdmin, IsTeacher

from .models import AbsenceExcuse, AttendanceRecord
from .serializers import (
    AbsenceExcuseSerializer,
    AbsenceStatsSerializer,
    AttendanceRecordSerializer,
    ExcuseReviewSerializer,
    ExcuseSubmitSerializer,
    JustifyAbsenceSerializer,
    MarkAttendanceSerializer,
)

# Reusable inline schemas
_AttendanceBulkResponseSchema = inline_serializer(
    "AttendanceBulkResponse",
    fields={
        "created": serializers.IntegerField(),
        "updated": serializers.IntegerField(),
        "errors": serializers.ListField(child=serializers.DictField()),
    },
)


# ---------------------------------------------------------------------------
# 1. MarkAttendanceView — teacher bulk-creates attendance records
# ---------------------------------------------------------------------------


class MarkAttendanceView(APIView):
    """
    POST /api/v1/attendance/mark/
    Teacher submits attendance for a class on a given date.
    Triggers parent notification for ABSENT students via Celery.
    """

    permission_classes = [permissions.IsAuthenticated, IsTeacher]

    @extend_schema(
        tags=["attendance"],
        summary="Mark attendance (teacher)",
        description=(
            "Teacher submits attendance records for a class on a given date. "
            "Parents of absent students are notified via push notifications."
        ),
        request=MarkAttendanceSerializer,
        responses={
            200: _AttendanceBulkResponseSchema,
            403: OpenApiResponse(description="Not a teacher or wrong school."),
            404: OpenApiResponse(description="Class not found."),
        },
    )
    def post(self, request):
        serializer = MarkAttendanceSerializer(data=request.data)
        serializer.is_valid(raise_exception=True)

        user = request.user
        data = serializer.validated_data

        # Resolve class
        try:
            klass = Class.objects.select_related("section", "level").get(
                pk=data["class_id"]
            )
        except Class.DoesNotExist:
            return Response(
                {"detail": "Class not found."},
                status=status.HTTP_404_NOT_FOUND,
            )

        # Verify class belongs to the teacher's school
        if klass.school_id != user.school_id:
            return Response(
                {"detail": "Class does not belong to your school."},
                status=status.HTTP_403_FORBIDDEN,
            )

        created = 0
        updated = 0
        errors = []
        absent_records = []

        for idx, item in enumerate(data["records"]):
            try:
                student = StudentProfile.objects.get(
                    pk=item["student_id"],
                    current_class=klass,
                )
            except StudentProfile.DoesNotExist:
                errors.append(
                    {"index": idx, "error": "Student not found in this class."}
                )
                continue

            record, was_created = AttendanceRecord.objects.update_or_create(
                student=student,
                date=data["date"],
                defaults={
                    "class_obj": klass,
                    "status": item["status"],
                    "note": item.get("note", ""),
                    "marked_by": user,
                    "school": user.school,
                },
            )
            if was_created:
                created += 1
            else:
                updated += 1

            if item["status"] == AttendanceRecord.Status.ABSENT:
                absent_records.append(record)

        # Trigger Celery notifications for absent students
        for record in absent_records:
            _notify_parents_absence(record)

        return Response(
            {"created": created, "updated": updated, "errors": errors},
            status=status.HTTP_200_OK,
        )


# ---------------------------------------------------------------------------
# 2. AttendanceListView — admin filterable list
# ---------------------------------------------------------------------------


class AttendanceListView(APIView):
    """
    GET /api/v1/attendance/
    Admin views attendance records with rich filtering, search, ordering
    and pagination.

    Query params
    ------------
    class_id, section_id, date, date_from, date_to, status,
    is_justified (true/false), student (UUID), teacher (UUID),
    search (student name substring), period (MORNING/AFTERNOON),
    ordering (-date default), page, page_size (default 50, max 200).
    """

    permission_classes = [permissions.IsAuthenticated, IsSchoolAdmin]

    DEFAULT_PAGE_SIZE = 50
    MAX_PAGE_SIZE = 200

    @extend_schema(
        tags=["attendance"],
        summary="List attendance records (admin)",
        description=(
            "List attendance records for the admin's school. "
            "Supports class_id, section_id, date, date_from, date_to, "
            "status, is_justified, student, teacher, search, period, "
            "ordering, page, page_size query params."
        ),
        responses={200: AttendanceRecordSerializer(many=True)},
    )
    def get(self, request):
        qs = AttendanceRecord.objects.filter(school=request.user.school)
        params = request.query_params

        # --- Filters ---
        if params.get("class_id"):
            qs = qs.filter(class_obj_id=params["class_id"])
        if params.get("section_id"):
            qs = qs.filter(class_obj__section_id=params["section_id"])
        if params.get("date"):
            qs = qs.filter(date=params["date"])
        if params.get("date_from"):
            qs = qs.filter(date__gte=params["date_from"])
        if params.get("date_to"):
            qs = qs.filter(date__lte=params["date_to"])
        if params.get("status"):
            qs = qs.filter(status=params["status"].upper())
        if params.get("is_justified") in ("true", "false"):
            qs = qs.filter(is_justified=params["is_justified"] == "true")
        if params.get("student"):
            qs = qs.filter(student_id=params["student"])
        if params.get("teacher"):
            qs = qs.filter(marked_by_id=params["teacher"])
        if params.get("period"):
            qs = qs.filter(period=params["period"].upper())
        if params.get("search"):
            qs = qs.filter(
                Q(student__user__first_name__icontains=params["search"])
                | Q(student__user__last_name__icontains=params["search"])
            )

        # --- Ordering ---
        ordering = params.get("ordering", "-date")
        allowed = {
            "date",
            "-date",
            "student__user__last_name",
            "-student__user__last_name",
        }
        if ordering in allowed:
            qs = qs.order_by(ordering, "-created_at")
        else:
            qs = qs.order_by("-date", "-created_at")

        # --- Eager loading ---
        qs = qs.select_related(
            "student__user", "class_obj", "marked_by", "justified_by"
        )

        # --- Pagination ---
        try:
            page = max(int(params.get("page", 1)), 1)
        except (ValueError, TypeError):
            page = 1
        try:
            page_size = min(
                max(int(params.get("page_size", self.DEFAULT_PAGE_SIZE)), 1),
                self.MAX_PAGE_SIZE,
            )
        except (ValueError, TypeError):
            page_size = self.DEFAULT_PAGE_SIZE

        total = qs.count()
        start = (page - 1) * page_size
        page_qs = qs[start : start + page_size]

        serializer = AttendanceRecordSerializer(page_qs, many=True)
        return Response(
            {
                "count": total,
                "page": page,
                "page_size": page_size,
                "results": serializer.data,
            }
        )


# ---------------------------------------------------------------------------
# 3. StudentAttendanceView — parent/student view
# ---------------------------------------------------------------------------


class StudentAttendanceView(APIView):
    """
    GET /api/v1/attendance/my/
    Parents see their children's attendance; students see their own.
    """

    permission_classes = [permissions.IsAuthenticated]

    @extend_schema(
        tags=["attendance"],
        summary="My attendance (student/parent)",
        description=(
            "Students see their own attendance history. "
            "Parents see attendance for all linked children. "
            "Optional **date** query param to filter by date."
        ),
        responses={
            200: AttendanceRecordSerializer(many=True),
            403: OpenApiResponse(description="Only students and parents can access."),
        },
    )
    def get(self, request):
        user = request.user

        if user.role == "STUDENT":
            profile = getattr(user, "student_profile", None)
            if not profile:
                return Response([])
            qs = AttendanceRecord.objects.filter(student=profile)

        elif user.role == "PARENT":
            parent_profile = getattr(user, "parent_profile", None)
            if not parent_profile:
                return Response([])
            children_ids = parent_profile.children.values_list("id", flat=True)
            qs = AttendanceRecord.objects.filter(student_id__in=children_ids)

        else:
            return Response(
                {"detail": "Not allowed."},
                status=status.HTTP_403_FORBIDDEN,
            )

        # Optional date filter
        date_param = request.query_params.get("date")
        if date_param:
            qs = qs.filter(date=date_param)

        qs = qs.select_related("student__user", "class_obj")
        serializer = AttendanceRecordSerializer(qs, many=True)
        return Response(serializer.data)


# ---------------------------------------------------------------------------
# 4. ExcuseSubmitView — parent submits an excuse
# ---------------------------------------------------------------------------


class ExcuseSubmitView(APIView):
    """
    POST /api/v1/attendance/excuses/
    Parent submits a justification for a student absence.
    """

    permission_classes = [permissions.IsAuthenticated, IsParent]

    @extend_schema(
        tags=["attendance"],
        summary="Submit absence excuse (parent)",
        description=(
            "Parent submits a justification for a child's absence. "
            "The parent must be linked to the student."
        ),
        request=ExcuseSubmitSerializer,
        responses={
            201: AbsenceExcuseSerializer,
            400: OpenApiResponse(description="Validation error."),
            403: OpenApiResponse(description="Not linked to this student."),
            404: OpenApiResponse(description="Attendance record not found."),
        },
    )
    def post(self, request):
        serializer = ExcuseSubmitSerializer(data=request.data)
        serializer.is_valid(raise_exception=True)

        data = serializer.validated_data

        record = get_object_or_404(
            AttendanceRecord,
            pk=data["attendance_record_id"],
            status=AttendanceRecord.Status.ABSENT,
        )

        # Verify parent is linked to this student
        parent_profile = getattr(request.user, "parent_profile", None)
        if (
            not parent_profile
            or not parent_profile.children.filter(pk=record.student_id).exists()
        ):
            return Response(
                {"detail": "You are not linked to this student."},
                status=status.HTTP_403_FORBIDDEN,
            )

        excuse = AbsenceExcuse.objects.create(
            attendance_record=record,
            submitted_by=request.user,
            justification_text=data["justification_text"],
            attachment=data.get("attachment"),
        )

        return Response(
            AbsenceExcuseSerializer(excuse).data,
            status=status.HTTP_201_CREATED,
        )


# ---------------------------------------------------------------------------
# 5. ExcuseReviewView — admin approves/rejects
# ---------------------------------------------------------------------------


class ExcuseReviewView(APIView):
    """
    PATCH /api/v1/attendance/excuses/<id>/review/
    Admin approves or rejects an absence excuse.
    """

    permission_classes = [permissions.IsAuthenticated, IsSchoolAdmin]

    @extend_schema(
        tags=["attendance"],
        summary="Review absence excuse (admin)",
        description=(
            "Admin approves or rejects a pending absence excuse. "
            "Notifies the parent of the result."
        ),
        request=ExcuseReviewSerializer,
        responses={
            200: AbsenceExcuseSerializer,
            400: OpenApiResponse(description="Excuse already reviewed."),
            403: OpenApiResponse(description="Not a school admin."),
            404: OpenApiResponse(description="Excuse not found."),
        },
    )
    def patch(self, request, pk):
        serializer = ExcuseReviewSerializer(data=request.data)
        serializer.is_valid(raise_exception=True)

        excuse = get_object_or_404(
            AbsenceExcuse,
            pk=pk,
            attendance_record__school=request.user.school,
        )

        if excuse.status != AbsenceExcuse.Status.PENDING:
            return Response(
                {"detail": f"Excuse already {excuse.status}."},
                status=status.HTTP_400_BAD_REQUEST,
            )

        new_status = serializer.validated_data["status"]
        excuse.status = new_status
        excuse.reviewed_by = request.user
        excuse.save(update_fields=["status", "reviewed_by"])

        # If approved, mark the attendance record as justified
        if new_status == AbsenceExcuse.Status.APPROVED:
            record = excuse.attendance_record
            record.justify(
                user=request.user,
                note=excuse.justification_text,
            )

        # Notify parent of result
        _notify_parent_excuse_reviewed(excuse)

        return Response(AbsenceExcuseSerializer(excuse).data)


# ---------------------------------------------------------------------------
# 6. AbsenceStatsView — dashboard stats for admin
# ---------------------------------------------------------------------------


class AbsenceStatsView(APIView):
    """
    GET /api/v1/attendance/stats/
    Dashboard statistics: today, week, month absence counts,
    at-risk students, teachers who haven't marked today.
    """

    permission_classes = [permissions.IsAuthenticated, IsSchoolAdmin]

    AT_RISK_THRESHOLD = 5  # unjustified absences this month

    @extend_schema(
        tags=["attendance"],
        summary="Absence dashboard stats (admin)",
        responses={200: AbsenceStatsSerializer},
    )
    def get(self, request):
        school = request.user.school
        today = timezone.localdate()
        week_start = today - timedelta(days=today.weekday())  # Monday
        month_start = today.replace(day=1)

        base = AttendanceRecord.objects.filter(
            school=school, status=AttendanceRecord.Status.ABSENT
        )

        today_count = base.filter(date=today).count()
        week_count = base.filter(date__gte=week_start).count()
        month_count = base.filter(date__gte=month_start).count()

        # At-risk students (≥ threshold unjustified absences this month)
        at_risk_qs = (
            base.filter(date__gte=month_start, is_justified=False)
            .values(
                "student__id", "student__user__first_name", "student__user__last_name"
            )
            .annotate(absence_count=Count("id"))
            .filter(absence_count__gte=self.AT_RISK_THRESHOLD)
            .order_by("-absence_count")[:20]
        )
        at_risk_students = [
            {
                "student_id": str(row["student__id"]),
                "student_name": f"{row['student__user__first_name']} {row['student__user__last_name']}".strip(),
                "absence_count": row["absence_count"],
            }
            for row in at_risk_qs
        ]

        # Teachers who haven't marked attendance today
        all_teacher_ids = set(
            User.objects.filter(
                school=school, role="TEACHER", is_active=True
            ).values_list("id", flat=True)
        )
        marked_teacher_ids = set(
            AttendanceRecord.objects.filter(school=school, date=today)
            .values_list("marked_by_id", flat=True)
            .distinct()
        )
        not_marked_ids = all_teacher_ids - marked_teacher_ids
        teachers_not_marked = list(
            User.objects.filter(id__in=not_marked_ids).values_list(
                "first_name", "last_name"
            )
        )
        teachers_not_marked = [f"{fn} {ln}".strip() for fn, ln in teachers_not_marked]

        data = {
            "today_count": today_count,
            "week_count": week_count,
            "month_count": month_count,
            "at_risk_students": at_risk_students,
            "teachers_not_marked": teachers_not_marked,
        }
        return Response(data)


# ---------------------------------------------------------------------------
# 7. JustifyAbsenceView — admin justifies an absence
# ---------------------------------------------------------------------------


class JustifyAbsenceView(APIView):
    """
    PATCH /api/v1/attendance/<id>/justify/
    Admin marks an absence as justified with a note.
    """

    permission_classes = [permissions.IsAuthenticated, IsSchoolAdmin]

    @extend_schema(
        tags=["attendance"],
        summary="Justify an absence (admin)",
        request=JustifyAbsenceSerializer,
        responses={
            200: AttendanceRecordSerializer,
            400: OpenApiResponse(description="Record is not an absence."),
            404: OpenApiResponse(description="Record not found."),
        },
    )
    def patch(self, request, pk):
        record = get_object_or_404(AttendanceRecord, pk=pk, school=request.user.school)
        if record.status != AttendanceRecord.Status.ABSENT:
            return Response(
                {"detail": "Only absences can be justified."},
                status=status.HTTP_400_BAD_REQUEST,
            )

        ser = JustifyAbsenceSerializer(data=request.data)
        ser.is_valid(raise_exception=True)

        record.justify(
            user=request.user,
            note=ser.validated_data["justification_note"],
        )
        return Response(AttendanceRecordSerializer(record).data)


# ---------------------------------------------------------------------------
# 8. CancelAbsenceView — admin deletes an attendance record
# ---------------------------------------------------------------------------


class CancelAbsenceView(APIView):
    """
    DELETE /api/v1/attendance/<id>/cancel/
    Admin deletes an attendance record (e.g. erroneous entry).
    """

    permission_classes = [permissions.IsAuthenticated, IsSchoolAdmin]

    @extend_schema(
        tags=["attendance"],
        summary="Cancel/delete attendance record (admin)",
        responses={
            204: OpenApiResponse(description="Deleted."),
            404: OpenApiResponse(description="Record not found."),
        },
    )
    def delete(self, request, pk):
        record = get_object_or_404(AttendanceRecord, pk=pk, school=request.user.school)
        record.delete()
        return Response(status=status.HTTP_204_NO_CONTENT)


# ---------------------------------------------------------------------------
# 9. AbsenceReportView — CSV export
# ---------------------------------------------------------------------------


class AbsenceReportView(APIView):
    """
    GET /api/v1/attendance/report/
    Export attendance data as CSV.
    Supports same filters as AttendanceListView.
    """

    permission_classes = [permissions.IsAuthenticated, IsSchoolAdmin]

    @extend_schema(
        tags=["attendance"],
        summary="Export attendance report (admin)",
        responses={200: OpenApiResponse(description="CSV file download.")},
    )
    def get(self, request):
        school = request.user.school
        params = request.query_params

        qs = AttendanceRecord.objects.filter(school=school)

        if params.get("class_id"):
            qs = qs.filter(class_obj_id=params["class_id"])
        if params.get("section_id"):
            qs = qs.filter(class_obj__section_id=params["section_id"])
        if params.get("date"):
            qs = qs.filter(date=params["date"])
        if params.get("date_from"):
            qs = qs.filter(date__gte=params["date_from"])
        if params.get("date_to"):
            qs = qs.filter(date__lte=params["date_to"])
        if params.get("status"):
            qs = qs.filter(status=params["status"].upper())
        if params.get("is_justified") in ("true", "false"):
            qs = qs.filter(is_justified=params["is_justified"] == "true")

        qs = qs.select_related("student__user", "class_obj", "marked_by").order_by(
            "-date", "student__user__last_name"
        )

        # Build CSV
        buf = io.StringIO()
        writer = csv.writer(buf)
        writer.writerow(
            [
                "Élève",
                "Classe",
                "Date",
                "Période",
                "Statut",
                "Justifié",
                "Note",
                "Marqué par",
            ]
        )
        for r in qs:
            writer.writerow(
                [
                    r.student.user.full_name if r.student and r.student.user else "",
                    str(r.class_obj) if r.class_obj else "",
                    str(r.date),
                    r.period or "",
                    r.get_status_display(),
                    "Oui" if r.is_justified else "Non",
                    r.note or "",
                    r.marked_by.full_name if r.marked_by else "",
                ]
            )

        response = HttpResponse(buf.getvalue(), content_type="text/csv; charset=utf-8")
        response["Content-Disposition"] = (
            f'attachment; filename="attendance_report_{timezone.localdate()}.csv"'
        )
        return response


# ===========================================================================
# Notification helpers (thin wrappers — actual sending via Celery tasks)
# ===========================================================================


def _notify_parents_absence(record):
    """Notify parents when their child is marked absent."""
    try:
        from apps.notifications.tasks import send_notification

        for parent_profile in record.student.parents.all():
            send_notification.delay(
                user_id=str(parent_profile.user_id),
                title="Absence Alert",
                body=(
                    f"{record.student.user.full_name} was marked absent "
                    f"on {record.date}."
                ),
            )
    except Exception:
        pass  # notifications are best-effort


def _notify_parent_excuse_reviewed(excuse):
    """Notify parent about excuse review result."""
    try:
        from apps.notifications.tasks import send_notification

        send_notification.delay(
            user_id=str(excuse.submitted_by_id),
            title=f"Excuse {excuse.status.lower()}",
            body=(
                f"Your absence excuse for "
                f"{excuse.attendance_record.student.user.full_name} "
                f"on {excuse.attendance_record.date} has been "
                f"{excuse.status.lower()}."
            ),
        )
    except Exception:
        pass


# ═══════════════════════════════════════════════════════════════════════════
# 10. ATTENDANCE REPORTS — monthly, annual, calendar, ranking, Excel
# ═══════════════════════════════════════════════════════════════════════════


class MonthlyAttendanceReportView(APIView):
    """
    GET /api/v1/attendance/reports/monthly/
    Query params: class_id, year, month
    Returns monthly attendance summary per student for a class.
    """

    permission_classes = [permissions.IsAuthenticated, IsSchoolAdmin]

    @extend_schema(tags=["attendance"], summary="Monthly attendance report by class")
    def get(self, request):
        import calendar
        from collections import defaultdict

        school = request.user.school
        class_id = request.query_params.get("class_id")
        year = int(request.query_params.get("year", timezone.localdate().year))
        month = int(request.query_params.get("month", timezone.localdate().month))

        if not class_id:
            return Response({"detail": "class_id requis."}, status=400)

        klass = get_object_or_404(Class, pk=class_id, section__school=school)

        _, days_in_month = calendar.monthrange(year, month)
        month_start = timezone.datetime(year, month, 1).date()
        month_end = timezone.datetime(year, month, days_in_month).date()

        students = StudentProfile.objects.filter(
            current_class=klass, is_deleted=False
        ).select_related("user").order_by("user__last_name", "user__first_name")

        records = AttendanceRecord.objects.filter(
            school=school,
            class_obj=klass,
            date__gte=month_start,
            date__lte=month_end,
        )

        # Build per-student summary
        student_data = defaultdict(lambda: {
            "absent": 0, "late": 0, "present": 0, "justified": 0, "days": {}
        })

        for r in records:
            sid = str(r.student_id)
            day = r.date.day
            student_data[sid]["days"][day] = r.status
            if r.status == AttendanceRecord.Status.ABSENT:
                student_data[sid]["absent"] += 1
                if r.is_justified:
                    student_data[sid]["justified"] += 1
            elif r.status == AttendanceRecord.Status.LATE:
                student_data[sid]["late"] += 1
            elif r.status == AttendanceRecord.Status.PRESENT:
                student_data[sid]["present"] += 1

        result = []
        for student in students:
            sid = str(student.pk)
            data = student_data.get(sid, {"absent": 0, "late": 0, "present": 0, "justified": 0, "days": {}})
            result.append({
                "student_id": sid,
                "name": student.user.get_full_name(),
                "absent": data["absent"],
                "late": data["late"],
                "present": data["present"],
                "justified": data["justified"],
                "days": data["days"],
            })

        return Response({
            "class_id": class_id,
            "class_name": klass.name,
            "year": year,
            "month": month,
            "days_in_month": days_in_month,
            "students": result,
        })


class StudentAttendanceCalendarView(APIView):
    """
    GET /api/v1/attendance/reports/calendar/
    Query params: student_id, academic_year_id (or year)
    Returns full calendar view of a student's attendance.
    """

    permission_classes = [permissions.IsAuthenticated]

    @extend_schema(tags=["attendance"], summary="Student attendance calendar")
    def get(self, request):
        student_id = request.query_params.get("student_id")
        if not student_id:
            return Response({"detail": "student_id requis."}, status=400)

        records = AttendanceRecord.objects.filter(
            student_id=student_id,
        ).order_by("date")

        year = request.query_params.get("year")
        if year:
            records = records.filter(date__year=int(year))

        calendar_data = {}
        for r in records:
            month_key = r.date.strftime("%Y-%m")
            if month_key not in calendar_data:
                calendar_data[month_key] = []
            calendar_data[month_key].append({
                "date": str(r.date),
                "day": r.date.day,
                "status": r.status,
                "period": r.period,
                "is_justified": r.is_justified,
            })

        # Summary
        total = records.count()
        absent = records.filter(status=AttendanceRecord.Status.ABSENT).count()
        late = records.filter(status=AttendanceRecord.Status.LATE).count()
        present = records.filter(status=AttendanceRecord.Status.PRESENT).count()
        justified = records.filter(is_justified=True).count()

        return Response({
            "student_id": student_id,
            "summary": {
                "total_records": total,
                "absent": absent,
                "late": late,
                "present": present,
                "justified": justified,
            },
            "calendar": calendar_data,
        })


class AnnualAttendanceReportView(APIView):
    """
    GET /api/v1/attendance/reports/annual/
    Query params: class_id (optional), year
    Returns annual attendance recap per month.
    """

    permission_classes = [permissions.IsAuthenticated, IsSchoolAdmin]

    @extend_schema(tags=["attendance"], summary="Annual attendance recap")
    def get(self, request):
        school = request.user.school
        year = int(request.query_params.get("year", timezone.localdate().year))
        class_id = request.query_params.get("class_id")

        base = AttendanceRecord.objects.filter(school=school, date__year=year)
        if class_id:
            base = base.filter(class_obj_id=class_id)

        monthly = (
            base.extra(select={"month": "EXTRACT(MONTH FROM date)"})
            .values("month")
            .annotate(
                total=Count("id"),
                absent=Count("id", filter=Q(status=AttendanceRecord.Status.ABSENT)),
                late=Count("id", filter=Q(status=AttendanceRecord.Status.LATE)),
                present=Count("id", filter=Q(status=AttendanceRecord.Status.PRESENT)),
                justified=Count("id", filter=Q(is_justified=True)),
            )
            .order_by("month")
        )

        return Response({
            "year": year,
            "class_id": class_id,
            "months": [
                {
                    "month": int(m["month"]),
                    "total": m["total"],
                    "absent": m["absent"],
                    "late": m["late"],
                    "present": m["present"],
                    "justified": m["justified"],
                    "absence_rate": round(m["absent"] / m["total"] * 100, 1) if m["total"] else 0,
                }
                for m in monthly
            ],
        })


class AttendanceRankingView(APIView):
    """
    GET /api/v1/attendance/reports/ranking/
    Query params: class_id (optional), date_from, date_to
    Returns students ranked by attendance (best → worst).
    """

    permission_classes = [permissions.IsAuthenticated, IsSchoolAdmin]

    @extend_schema(tags=["attendance"], summary="Student attendance ranking")
    def get(self, request):
        school = request.user.school
        class_id = request.query_params.get("class_id")
        date_from = request.query_params.get("date_from")
        date_to = request.query_params.get("date_to")

        base = AttendanceRecord.objects.filter(school=school)
        if class_id:
            base = base.filter(class_obj_id=class_id)
        if date_from:
            base = base.filter(date__gte=date_from)
        if date_to:
            base = base.filter(date__lte=date_to)

        ranking = (
            base.values(
                "student__id",
                "student__user__first_name",
                "student__user__last_name",
            )
            .annotate(
                total=Count("id"),
                absent=Count("id", filter=Q(status=AttendanceRecord.Status.ABSENT)),
                late=Count("id", filter=Q(status=AttendanceRecord.Status.LATE)),
                present=Count("id", filter=Q(status=AttendanceRecord.Status.PRESENT)),
            )
            .order_by("absent", "late")
        )

        result = []
        for idx, r in enumerate(ranking, 1):
            total = r["total"] or 1
            result.append({
                "rank": idx,
                "student_id": str(r["student__id"]),
                "name": f"{r['student__user__first_name']} {r['student__user__last_name']}",
                "total": r["total"],
                "present": r["present"],
                "absent": r["absent"],
                "late": r["late"],
                "attendance_rate": round(r["present"] / total * 100, 1),
            })

        return Response(result)


class AttendanceExcelExportView(APIView):
    """
    GET /api/v1/attendance/reports/excel/
    Query params: class_id, date_from, date_to
    Exports attendance data as Excel file.
    """

    permission_classes = [permissions.IsAuthenticated, IsSchoolAdmin]

    @extend_schema(tags=["attendance"], summary="Export attendance to Excel")
    def get(self, request):
        import openpyxl
        from openpyxl.styles import Alignment, Font, PatternFill

        school = request.user.school
        class_id = request.query_params.get("class_id")
        date_from = request.query_params.get("date_from")
        date_to = request.query_params.get("date_to")

        base = AttendanceRecord.objects.filter(school=school)
        if class_id:
            base = base.filter(class_obj_id=class_id)
        if date_from:
            base = base.filter(date__gte=date_from)
        if date_to:
            base = base.filter(date__lte=date_to)

        base = base.select_related(
            "student__user", "class_obj", "marked_by"
        ).order_by("date", "student__user__last_name")

        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "Attendance Report"

        # Header style
        header_fill = PatternFill(start_color="1F4E79", end_color="1F4E79", fill_type="solid")
        header_font = Font(bold=True, color="FFFFFF")

        headers = ["Élève", "Classe", "Date", "Période", "Statut", "Justifié", "Note"]
        for col, h in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col, value=h)
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = Alignment(horizontal="center")

        status_fills = {
            "ABSENT": PatternFill(start_color="FF6B6B", end_color="FF6B6B", fill_type="solid"),
            "LATE": PatternFill(start_color="FFD93D", end_color="FFD93D", fill_type="solid"),
            "PRESENT": PatternFill(start_color="6BCB77", end_color="6BCB77", fill_type="solid"),
        }

        for row_idx, r in enumerate(base[:5000], 2):
            ws.cell(row=row_idx, column=1, value=r.student.user.get_full_name())
            ws.cell(row=row_idx, column=2, value=str(r.class_obj) if r.class_obj else "")
            ws.cell(row=row_idx, column=3, value=str(r.date))
            ws.cell(row=row_idx, column=4, value=r.period or "")
            status_cell = ws.cell(row=row_idx, column=5, value=r.get_status_display())
            if r.status in status_fills:
                status_cell.fill = status_fills[r.status]
            ws.cell(row=row_idx, column=6, value="Oui" if r.is_justified else "Non")
            ws.cell(row=row_idx, column=7, value=r.note or "")

        # Auto-width
        for col in ws.columns:
            max_len = max(len(str(c.value or "")) for c in col)
            ws.column_dimensions[col[0].column_letter].width = max(max_len + 2, 10)

        response = HttpResponse(
            content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        )
        response["Content-Disposition"] = (
            f'attachment; filename="attendance_report_{timezone.localdate()}.xlsx"'
        )
        wb.save(response)
        return response
