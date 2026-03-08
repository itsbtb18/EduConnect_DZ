"""
╔══════════════════════════════════════════════════════════════════════════╗
║  Grades Views — complete REST API                                      ║
║                                                                        ║
║  ExamType           CRUD + percentage validation                       ║
║  Grade              bulk-enter / list / publish / correct              ║
║  SubjectAverage     list / recalculate / override / publish            ║
║  TrimesterAverage   list / recalculate / override / publish / lock     ║
║  GradeAppeal        create / list / respond / pending-count            ║
╚══════════════════════════════════════════════════════════════════════════╝
"""

import logging
from decimal import Decimal

from django.db.models import Q
from django.shortcuts import get_object_or_404
from django.utils import timezone
from drf_spectacular.utils import extend_schema
from rest_framework import permissions, status, viewsets
from rest_framework.decorators import action
from rest_framework.response import Response
from rest_framework.views import APIView

from apps.academics.models import StudentProfile, TeacherAssignment
from apps.accounts.models import User

from .models import (
    AnnualAverage,
    ExamType,
    Grade,
    GradeAppeal,
    GradeAuditLog,
    ReportCard,
    ReportCardTemplate,
    SubjectAverage,
    TrimesterAverage,
    TrimesterConfig,
)
from .audit import log_grade_action
from .permissions import (
    IsAdminOnly,
    IsAdminOrTeacherOfClassroom,
    IsNotLocked,
    IsTeacherOfClassroom,
)
from .serializers import (
    AnnualAverageSerializer,
    CSVImportConfirmSerializer,
    CSVImportPreviewSerializer,
    ExamTypeCreateSerializer,
    ExamTypeSerializer,
    GenerateClassReportCardsSerializer,
    GenerateSchoolReportCardsSerializer,
    GradeAppealCreateSerializer,
    GradeAppealRespondSerializer,
    GradeAppealSerializer,
    GradeAuditLogSerializer,
    GradeBulkEnterSerializer,
    GradeCorrectSerializer,
    GradePublishSerializer,
    GradeReturnSerializer,
    GradeSerializer,
    GradeSubmitSerializer,
    ReportCardSerializer,
    ReportCardTemplateSerializer,
    SubjectAverageOverrideSerializer,
    SubjectAveragePublishSerializer,
    SubjectAverageRecalcSerializer,
    SubjectAverageSerializer,
    TrimesterAverageSerializer,
    TrimesterConfigSerializer,
    TrimesterLockSerializer,
    TrimesterOverrideSerializer,
    TrimesterPublishSerializer,
    TrimesterRecalcSerializer,
    TrimesterUnlockSerializer,
)

logger = logging.getLogger(__name__)


def _resolve_school(request):
    """Resolve school for the current user."""
    user = request.user
    if user.school_id:
        return user.school
    from apps.schools.models import School

    school_id = request.query_params.get("school") or request.data.get("school")
    if school_id:
        return School.objects.filter(pk=school_id, is_deleted=False).first()
    schools = School.objects.filter(is_deleted=False)
    if schools.count() == 1:
        return schools.first()
    return None


def _get_active_academic_year(school):
    """Return the current academic year for a school."""
    from apps.schools.models import AcademicYear

    return AcademicYear.objects.filter(school=school, is_current=True).first()


# ═══════════════════════════════════════════════════════════════════════════
# 1. EXAM TYPE — CRUD
# ═══════════════════════════════════════════════════════════════════════════


class ExamTypeViewSet(viewsets.ModelViewSet):
    """
    CRUD for ExamType.
    - POST/PUT/DELETE: IsAdminOrTeacherOfClassroom + IsNotLocked
    - GET: IsAdminOrTeacherOfClassroom
    """

    permission_classes = [permissions.IsAuthenticated, IsAdminOrTeacherOfClassroom]

    def get_serializer_class(self):
        if self.action in ("create", "update", "partial_update"):
            return ExamTypeCreateSerializer
        return ExamTypeSerializer

    def get_queryset(self):
        school = _resolve_school(self.request)
        if not school:
            return ExamType.objects.none()

        qs = ExamType.objects.filter(
            classroom__section__school=school,
        ).select_related("subject", "classroom", "academic_year")

        # Filters
        classroom_id = self.request.query_params.get("classroom_id")
        subject_id = self.request.query_params.get("subject_id")
        trimester = self.request.query_params.get("trimester")

        if classroom_id:
            qs = qs.filter(classroom_id=classroom_id)
        if subject_id:
            qs = qs.filter(subject_id=subject_id)
        if trimester:
            qs = qs.filter(trimester=int(trimester))

        return qs

    def perform_create(self, serializer):
        serializer.save(created_by=self.request.user)

    def get_permissions(self):
        if self.action in ("update", "partial_update", "destroy"):
            return [
                permissions.IsAuthenticated(),
                IsAdminOrTeacherOfClassroom(),
                IsNotLocked(),
            ]
        return super().get_permissions()

    @extend_schema(tags=["grades"], summary="Delete exam type")
    def destroy(self, request, *args, **kwargs):
        """Vérifier qu'aucune note n'existe avant suppression."""
        instance = self.get_object()
        grade_count = Grade.objects.filter(exam_type=instance).count()
        if grade_count > 0:
            return Response(
                {
                    "detail": (
                        f"Impossible de supprimer : {grade_count} note(s) existent "
                        "pour ce type d'examen. Supprimez les notes d'abord."
                    ),
                    "grade_count": grade_count,
                },
                status=status.HTTP_409_CONFLICT,
            )
        return super().destroy(request, *args, **kwargs)


# ═══════════════════════════════════════════════════════════════════════════
# 2. GRADE — bulk-enter / list / publish / correct
# ═══════════════════════════════════════════════════════════════════════════


class GradeBulkEnterView(APIView):
    """
    POST /api/grades/bulk-enter/
    Body: { exam_type_id, grades: [{student_id, score, is_absent}] }
    Permission: IsTeacherOfClassroom + IsNotLocked
    """

    permission_classes = [
        permissions.IsAuthenticated,
        IsTeacherOfClassroom,
        IsNotLocked,
    ]

    @extend_schema(
        tags=["grades"],
        summary="Bulk enter grades for an exam",
        request=GradeBulkEnterSerializer,
    )
    def post(self, request):
        ser = GradeBulkEnterSerializer(data=request.data)
        ser.is_valid(raise_exception=True)

        exam_type = get_object_or_404(ExamType, pk=ser.validated_data["exam_type_id"])
        user = request.user
        saved = 0
        errors = []

        for idx, item in enumerate(ser.validated_data["grades"]):
            try:
                student = StudentProfile.objects.get(
                    pk=item["student_id"],
                    is_deleted=False,
                )
            except StudentProfile.DoesNotExist:
                errors.append({"index": idx, "error": "Élève introuvable."})
                continue

            score = item.get("score")
            is_absent = item.get("is_absent", False)

            if score is not None and score > exam_type.max_score:
                errors.append(
                    {
                        "index": idx,
                        "error": f"Note {score} dépasse le barème {exam_type.max_score}.",
                    }
                )
                continue

            grade, _ = Grade.objects.update_or_create(
                student=student,
                exam_type=exam_type,
                defaults={
                    "score": score,
                    "is_absent": is_absent,
                    "entered_by": user,
                },
            )
            saved += 1
            # Audit
            log_grade_action(
                action="GRADE_ENTERED",
                instance=grade,
                performed_by=user,
                request=request,
                student=student,
                new_value=score,
                subject_name=str(exam_type.subject),
                exam_name=exam_type.name,
                trimester=exam_type.trimester,
            )
            # Note: post_save signal on Grade will trigger async_recalculate_cascade

        return Response(
            {"saved": saved, "errors": errors},
            status=status.HTTP_200_OK,
        )


class GradeListView(APIView):
    """
    GET /api/grades/?classroom_id=X&exam_type_id=Y
    """

    permission_classes = [permissions.IsAuthenticated, IsAdminOrTeacherOfClassroom]

    @extend_schema(tags=["grades"], summary="List grades")
    def get(self, request):
        qs = Grade.objects.select_related(
            "student__user",
            "exam_type__subject",
            "exam_type__classroom",
        )

        classroom_id = request.query_params.get("classroom_id")
        exam_type_id = request.query_params.get("exam_type_id")
        student_id = request.query_params.get("student_id")
        trimester = request.query_params.get("trimester")

        if classroom_id:
            qs = qs.filter(exam_type__classroom_id=classroom_id)
        if exam_type_id:
            qs = qs.filter(exam_type_id=exam_type_id)
        if student_id:
            qs = qs.filter(student_id=student_id)
        if trimester:
            qs = qs.filter(exam_type__trimester=int(trimester))

        # Tenant isolation: only grades from user's school
        school = _resolve_school(request)
        if school:
            qs = qs.filter(exam_type__classroom__section__school=school)

        serializer = GradeSerializer(qs, many=True)
        return Response(serializer.data)


class GradePublishView(APIView):
    """
    POST /api/grades/publish/
    Body: { exam_type_id }
    Permission: IsAdminOnly (admin approves submitted grades → PUBLISHED)

    Workflow: SUBMITTED → PUBLISHED
    Sets is_published=True, sends notifications to students/parents.
    """

    permission_classes = [permissions.IsAuthenticated, IsAdminOnly]

    @extend_schema(
        tags=["grades"],
        summary="Publish grades (admin approval)",
        request=GradePublishSerializer,
    )
    def post(self, request):
        ser = GradePublishSerializer(data=request.data)
        ser.is_valid(raise_exception=True)

        from .services import InvalidTransitionError, publish_grades

        try:
            result = publish_grades(
                exam_type_id=ser.validated_data["exam_type_id"],
                admin_user=request.user,
            )
        except InvalidTransitionError as e:
            return Response(
                {"detail": str(e)},
                status=status.HTTP_400_BAD_REQUEST,
            )
        except Exception as e:
            return Response(
                {"detail": str(e)},
                status=status.HTTP_400_BAD_REQUEST,
            )

        # Audit
        exam_type = get_object_or_404(ExamType, pk=ser.validated_data["exam_type_id"])
        grades = Grade.objects.filter(
            exam_type=exam_type,
            status=Grade.Status.PUBLISHED,
        ).select_related("student")
        for g in grades:
            log_grade_action(
                action="GRADE_PUBLISHED",
                instance=g,
                performed_by=request.user,
                request=request,
                student=g.student,
                new_value=g.score,
                subject_name=str(exam_type.subject),
                exam_name=exam_type.name,
                trimester=exam_type.trimester,
            )

        return Response(result)


class GradeCorrectView(APIView):
    """
    POST /api/grades/{id}/correct/
    Body: { new_score, reason }
    Permission: IsAdminOrTeacherOfClassroom + IsNotLocked
    Modifier la note + recalcul en cascade + log.
    """

    permission_classes = [
        permissions.IsAuthenticated,
        IsAdminOrTeacherOfClassroom,
        IsNotLocked,
    ]

    @extend_schema(
        tags=["grades"],
        summary="Correct a grade",
        request=GradeCorrectSerializer,
    )
    def post(self, request, pk):
        ser = GradeCorrectSerializer(data=request.data)
        ser.is_valid(raise_exception=True)

        grade = get_object_or_404(
            Grade.objects.select_related("exam_type"),
            pk=pk,
        )

        old_score = grade.score
        new_score = ser.validated_data["new_score"]
        reason = ser.validated_data["reason"]

        # Validate against max_score
        if new_score > grade.exam_type.max_score:
            return Response(
                {
                    "detail": f"La note {new_score} dépasse le barème {grade.exam_type.max_score}."
                },
                status=status.HTTP_400_BAD_REQUEST,
            )

        # Update grade
        grade.score = new_score
        grade.is_absent = False
        grade.entered_by = request.user
        grade.save()
        # post_save signal triggers cascade

        # Audit
        log_grade_action(
            action="GRADE_CORRECTED",
            instance=grade,
            performed_by=request.user,
            request=request,
            student=grade.student,
            old_value=old_score,
            new_value=new_score,
            reason=reason,
            subject_name=str(grade.exam_type.subject),
            exam_name=grade.exam_type.name,
            trimester=grade.exam_type.trimester,
        )

        logger.info(
            "GRADE_CORRECT: grade=%s old=%s new=%s reason='%s' by=%s",
            pk,
            old_score,
            new_score,
            reason,
            request.user.pk,
        )

        return Response(
            {
                "id": str(grade.pk),
                "old_score": str(old_score) if old_score is not None else None,
                "new_score": str(new_score),
                "reason": reason,
            }
        )


# ═══════════════════════════════════════════════════════════════════════════
# 3. SUBJECT AVERAGE — list / recalculate / override / publish
# ═══════════════════════════════════════════════════════════════════════════


class SubjectAverageListView(APIView):
    """
    GET /api/grades/subject-averages/?classroom_id=X&trimester=1
    """

    permission_classes = [permissions.IsAuthenticated]

    @extend_schema(tags=["grades"], summary="List subject averages")
    def get(self, request):
        qs = SubjectAverage.objects.select_related(
            "student__user",
            "subject",
            "classroom",
        )

        classroom_id = request.query_params.get("classroom_id")
        trimester = request.query_params.get("trimester")
        student_id = request.query_params.get("student_id")
        subject_id = request.query_params.get("subject_id")

        if classroom_id:
            qs = qs.filter(classroom_id=classroom_id)
        if trimester:
            qs = qs.filter(trimester=int(trimester))
        if student_id:
            qs = qs.filter(student_id=student_id)
        if subject_id:
            qs = qs.filter(subject_id=subject_id)

        # Tenant isolation
        school = _resolve_school(request)
        if school:
            qs = qs.filter(classroom__section__school=school)

        serializer = SubjectAverageSerializer(qs, many=True)
        return Response(serializer.data)


class SubjectAverageRecalculateView(APIView):
    """
    POST /api/grades/subject-averages/recalculate/
    Body: { classroom_id, subject_id, trimester }
    Forcer le recalcul pour toute la classe.
    """

    permission_classes = [
        permissions.IsAuthenticated,
        IsAdminOrTeacherOfClassroom,
    ]

    @extend_schema(
        tags=["grades"],
        summary="Force subject average recalculation",
        request=SubjectAverageRecalcSerializer,
    )
    def post(self, request):
        ser = SubjectAverageRecalcSerializer(data=request.data)
        ser.is_valid(raise_exception=True)

        from apps.academics.models import Class

        classroom = get_object_or_404(Class, pk=ser.validated_data["classroom_id"])
        subject_id = ser.validated_data["subject_id"]
        trimester = ser.validated_data["trimester"]
        academic_year = _get_active_academic_year(_resolve_school(request))

        if not academic_year:
            return Response(
                {"detail": "Aucune année scolaire active trouvée."},
                status=status.HTTP_400_BAD_REQUEST,
            )

        from .services import LockedException, calculate_subject_average

        students = StudentProfile.objects.filter(
            current_class=classroom,
            is_deleted=False,
        )

        recalculated = 0
        locked_skipped = 0
        for student in students:
            try:
                result = calculate_subject_average(
                    student.pk,
                    subject_id,
                    classroom.pk,
                    academic_year.pk,
                    trimester,
                    save=True,
                )
                if result is not None:
                    recalculated += 1
            except LockedException:
                locked_skipped += 1

        return Response(
            {
                "recalculated": recalculated,
                "locked_skipped": locked_skipped,
            }
        )


class SubjectAverageOverrideView(APIView):
    """
    POST /api/grades/subject-averages/override/
    Body: { subject_average_id, new_value, reason }
    Permission: IsAdminOnly
    """

    permission_classes = [permissions.IsAuthenticated, IsAdminOnly]

    @extend_schema(
        tags=["grades"],
        summary="Override subject average (admin)",
        request=SubjectAverageOverrideSerializer,
    )
    def post(self, request):
        ser = SubjectAverageOverrideSerializer(data=request.data)
        ser.is_valid(raise_exception=True)

        from .services import (
            LockedException,
            PermissionDeniedException,
            admin_override_average,
        )

        sa = get_object_or_404(
            SubjectAverage, pk=ser.validated_data["subject_average_id"]
        )
        old_val = sa.effective_average

        try:
            result = admin_override_average(
                override_type="subject_average",
                object_id=ser.validated_data["subject_average_id"],
                new_value=ser.validated_data["new_value"],
                admin_user=request.user,
            )
        except LockedException as e:
            return Response({"detail": str(e)}, status=status.HTTP_423_LOCKED)
        except PermissionDeniedException as e:
            return Response({"detail": str(e)}, status=status.HTTP_403_FORBIDDEN)

        # Audit
        log_grade_action(
            action="AVERAGE_OVERRIDDEN",
            instance=sa,
            performed_by=request.user,
            request=request,
            student=sa.student,
            old_value=old_val,
            new_value=ser.validated_data["new_value"],
            reason=ser.validated_data["reason"],
            subject_name=str(sa.subject),
            trimester=sa.trimester,
        )

        logger.info(
            "SubjectAverage override — reason='%s' by=%s",
            ser.validated_data["reason"],
            request.user.pk,
        )
        return Response(result)


class SubjectAveragePublishView(APIView):
    """
    POST /api/grades/subject-averages/publish/
    Body: { classroom_id, subject_id, trimester }
    Permission: IsAdminOnly
    """

    permission_classes = [permissions.IsAuthenticated, IsAdminOnly]

    @extend_schema(
        tags=["grades"],
        summary="Publish subject averages (admin only)",
        request=SubjectAveragePublishSerializer,
    )
    def post(self, request):
        ser = SubjectAveragePublishSerializer(data=request.data)
        ser.is_valid(raise_exception=True)

        now = timezone.now()
        count = SubjectAverage.objects.filter(
            classroom_id=ser.validated_data["classroom_id"],
            subject_id=ser.validated_data["subject_id"],
            trimester=ser.validated_data["trimester"],
            is_published=False,
        ).update(
            is_published=True,
            published_at=now,
            published_by=request.user,
        )

        return Response({"published_count": count})


# ═══════════════════════════════════════════════════════════════════════════
# 4. TRIMESTER AVERAGE — list / recalculate / override / publish / lock
# ═══════════════════════════════════════════════════════════════════════════


class TrimesterAverageListView(APIView):
    """
    GET /api/grades/trimester-averages/?classroom_id=X&trimester=1
    """

    permission_classes = [permissions.IsAuthenticated]

    @extend_schema(tags=["grades"], summary="List trimester averages")
    def get(self, request):
        qs = TrimesterAverage.objects.select_related(
            "student__user",
            "classroom",
        )

        classroom_id = request.query_params.get("classroom_id")
        trimester = request.query_params.get("trimester")
        student_id = request.query_params.get("student_id")

        if classroom_id:
            qs = qs.filter(classroom_id=classroom_id)
        if trimester:
            qs = qs.filter(trimester=int(trimester))
        if student_id:
            qs = qs.filter(student_id=student_id)

        school = _resolve_school(request)
        if school:
            qs = qs.filter(classroom__section__school=school)

        serializer = TrimesterAverageSerializer(qs, many=True)
        return Response(serializer.data)


class TrimesterRecalculateView(APIView):
    """
    POST /api/grades/trimester-averages/recalculate/
    Body: { classroom_id, trimester }
    Permission: IsAdminOnly
    """

    permission_classes = [permissions.IsAuthenticated, IsAdminOnly]

    @extend_schema(
        tags=["grades"],
        summary="Recalculate trimester averages + rankings",
        request=TrimesterRecalcSerializer,
    )
    def post(self, request):
        ser = TrimesterRecalcSerializer(data=request.data)
        ser.is_valid(raise_exception=True)

        from .services import recalculate_classroom_trimester

        school = _resolve_school(request)
        academic_year = _get_active_academic_year(school) if school else None
        if not academic_year:
            return Response(
                {"detail": "Aucune année scolaire active trouvée."},
                status=status.HTTP_400_BAD_REQUEST,
            )

        summary = recalculate_classroom_trimester(
            ser.validated_data["classroom_id"],
            academic_year.pk,
            ser.validated_data["trimester"],
        )
        return Response(summary)


class TrimesterOverrideView(APIView):
    """
    POST /api/grades/trimester-averages/override/
    Body: { trimester_average_id, new_value, reason }
    Permission: IsAdminOnly
    """

    permission_classes = [permissions.IsAuthenticated, IsAdminOnly]

    @extend_schema(
        tags=["grades"],
        summary="Override trimester average (admin)",
        request=TrimesterOverrideSerializer,
    )
    def post(self, request):
        ser = TrimesterOverrideSerializer(data=request.data)
        ser.is_valid(raise_exception=True)

        from .services import (
            LockedException,
            PermissionDeniedException,
            admin_override_average,
        )

        ta = get_object_or_404(
            TrimesterAverage, pk=ser.validated_data["trimester_average_id"]
        )
        old_val = ta.effective_average

        try:
            result = admin_override_average(
                override_type="trimester_average",
                object_id=ser.validated_data["trimester_average_id"],
                new_value=ser.validated_data["new_value"],
                admin_user=request.user,
            )
        except LockedException as e:
            return Response({"detail": str(e)}, status=status.HTTP_423_LOCKED)
        except PermissionDeniedException as e:
            return Response({"detail": str(e)}, status=status.HTTP_403_FORBIDDEN)

        # Audit
        log_grade_action(
            action="AVERAGE_OVERRIDDEN",
            instance=ta,
            performed_by=request.user,
            request=request,
            student=ta.student,
            old_value=old_val,
            new_value=ser.validated_data["new_value"],
            reason=ser.validated_data["reason"],
            trimester=ta.trimester,
        )

        logger.info(
            "TrimesterAverage override — reason='%s' by=%s",
            ser.validated_data["reason"],
            request.user.pk,
        )
        return Response(result)


class TrimesterPublishView(APIView):
    """
    POST /api/grades/trimester-averages/publish/
    Body: { classroom_id, trimester }
    Permission: IsAdminOnly
    """

    permission_classes = [permissions.IsAuthenticated, IsAdminOnly]

    @extend_schema(
        tags=["grades"],
        summary="Publish trimester averages (admin only)",
        request=TrimesterPublishSerializer,
    )
    def post(self, request):
        ser = TrimesterPublishSerializer(data=request.data)
        ser.is_valid(raise_exception=True)

        now = timezone.now()
        count = TrimesterAverage.objects.filter(
            classroom_id=ser.validated_data["classroom_id"],
            trimester=ser.validated_data["trimester"],
            is_published=False,
        ).update(
            is_published=True,
            published_at=now,
            published_by=request.user,
        )

        return Response({"published_count": count})


class TrimesterLockView(APIView):
    """
    POST /api/grades/trimester-averages/lock/
    Body: { classroom_id, trimester }
    Permission: IsAdminOnly
    """

    permission_classes = [permissions.IsAuthenticated, IsAdminOnly]

    @extend_schema(
        tags=["grades"],
        summary="Lock trimester",
        request=TrimesterLockSerializer,
    )
    def post(self, request):
        ser = TrimesterLockSerializer(data=request.data)
        ser.is_valid(raise_exception=True)

        from .services import LockedException, PermissionDeniedException, lock_trimester

        school = _resolve_school(request)
        academic_year = _get_active_academic_year(school) if school else None
        if not academic_year:
            return Response(
                {"detail": "Aucune année scolaire active trouvée."},
                status=status.HTTP_400_BAD_REQUEST,
            )

        try:
            result = lock_trimester(
                classroom_id=ser.validated_data["classroom_id"],
                academic_year_id=academic_year.pk,
                trimester=ser.validated_data["trimester"],
                director_user=request.user,
                check_completeness=True,
            )
        except PermissionDeniedException as e:
            return Response({"detail": str(e)}, status=status.HTTP_403_FORBIDDEN)
        except ValueError as e:
            return Response({"detail": str(e)}, status=status.HTTP_400_BAD_REQUEST)

        # Audit — log for each locked trimester average
        locked_avgs = TrimesterAverage.objects.filter(
            classroom_id=ser.validated_data["classroom_id"],
            trimester=ser.validated_data["trimester"],
            is_locked=True,
        ).select_related("student")
        for ta in locked_avgs:
            log_grade_action(
                action="TRIMESTER_LOCKED",
                instance=ta,
                performed_by=request.user,
                request=request,
                student=ta.student,
                trimester=ser.validated_data["trimester"],
            )

        return Response(result)


class TrimesterUnlockView(APIView):
    """
    POST /api/grades/trimester-averages/unlock/
    Body: { classroom_id, trimester, reason }
    Permission: IsAdminOnly
    """

    permission_classes = [permissions.IsAuthenticated, IsAdminOnly]

    @extend_schema(
        tags=["grades"],
        summary="Unlock trimester",
        request=TrimesterUnlockSerializer,
    )
    def post(self, request):
        ser = TrimesterUnlockSerializer(data=request.data)
        ser.is_valid(raise_exception=True)

        from .services import PermissionDeniedException, unlock_trimester

        school = _resolve_school(request)
        academic_year = _get_active_academic_year(school) if school else None
        if not academic_year:
            return Response(
                {"detail": "Aucune année scolaire active trouvée."},
                status=status.HTTP_400_BAD_REQUEST,
            )

        try:
            result = unlock_trimester(
                classroom_id=ser.validated_data["classroom_id"],
                academic_year_id=academic_year.pk,
                trimester=ser.validated_data["trimester"],
                director_user=request.user,
                reason=ser.validated_data["reason"],
            )
        except PermissionDeniedException as e:
            return Response({"detail": str(e)}, status=status.HTTP_403_FORBIDDEN)
        except ValueError as e:
            return Response({"detail": str(e)}, status=status.HTTP_400_BAD_REQUEST)

        # Audit — log for each unlocked trimester average
        unlocked_avgs = TrimesterAverage.objects.filter(
            classroom_id=ser.validated_data["classroom_id"],
            trimester=ser.validated_data["trimester"],
            is_locked=False,
        ).select_related("student")
        for ta in unlocked_avgs:
            log_grade_action(
                action="TRIMESTER_UNLOCKED",
                instance=ta,
                performed_by=request.user,
                request=request,
                student=ta.student,
                trimester=ser.validated_data["trimester"],
                reason=ser.validated_data["reason"],
            )

        return Response(result)


# ═══════════════════════════════════════════════════════════════════════════
# 5. GRADE APPEAL — create / list / respond / pending-count
# ═══════════════════════════════════════════════════════════════════════════


class GradeAppealCreateView(APIView):
    """
    POST /api/grades/appeals/
    Permission: IsAuthenticated (student from their app)
    """

    permission_classes = [permissions.IsAuthenticated]

    @extend_schema(
        tags=["grades"],
        summary="Create a grade appeal",
        request=GradeAppealCreateSerializer,
    )
    def post(self, request):
        ser = GradeAppealCreateSerializer(data=request.data)
        ser.is_valid(raise_exception=True)
        data = ser.validated_data

        # Resolve student profile
        try:
            student = request.user.student_profile
        except StudentProfile.DoesNotExist:
            return Response(
                {"detail": "Profil étudiant introuvable."},
                status=status.HTTP_400_BAD_REQUEST,
            )

        appeal_type = data["appeal_type"]
        appeal = GradeAppeal(
            student=student,
            appeal_type=appeal_type,
            reason=data["reason"],
            student_comment=data.get("student_comment", ""),
            status=GradeAppeal.Status.PENDING,
        )

        # Resolve FK + auto-assign teacher or admin
        if appeal_type == GradeAppeal.AppealType.EXAM_GRADE:
            grade = get_object_or_404(Grade, pk=data["grade_id"])
            appeal.grade = grade
            appeal.original_value = grade.score
            # Assign to the teacher of that subject in that class
            teacher = self._find_teacher(
                grade.exam_type.classroom,
                grade.exam_type.subject,
            )
            if teacher:
                appeal.assigned_to_teacher = teacher

        elif appeal_type == GradeAppeal.AppealType.SUBJECT_AVERAGE:
            sa = get_object_or_404(SubjectAverage, pk=data["subject_average_id"])
            appeal.subject_average = sa
            appeal.original_value = sa.effective_average
            teacher = self._find_teacher(sa.classroom, sa.subject)
            if teacher:
                appeal.assigned_to_teacher = teacher

        elif appeal_type == GradeAppeal.AppealType.TRIMESTER_AVERAGE:
            ta = get_object_or_404(TrimesterAverage, pk=data["trimester_average_id"])
            appeal.trimester_average = ta
            appeal.original_value = ta.effective_average
            appeal.assigned_to_admin = True

        elif appeal_type == GradeAppeal.AppealType.ANNUAL_AVERAGE:
            appeal.assigned_to_admin = True

        appeal.save()
        # Signal post_save triggers notify_appeal_assigned

        # Audit
        log_grade_action(
            action="APPEAL_CREATED",
            instance=appeal,
            performed_by=request.user,
            request=request,
            student=student,
            old_value=appeal.original_value,
            reason=data["reason"],
        )

        return Response(
            GradeAppealSerializer(appeal).data,
            status=status.HTTP_201_CREATED,
        )

    @staticmethod
    def _find_teacher(classroom, subject):
        """Find the TeacherProfile assigned to this subject/classroom."""
        from apps.academics.models import TeacherProfile

        assignment = (
            TeacherAssignment.objects.filter(
                assigned_class=classroom,
                subject=subject,
            )
            .select_related("teacher__teacher_profile")
            .first()
        )

        if assignment:
            try:
                return assignment.teacher.teacher_profile
            except TeacherProfile.DoesNotExist:
                pass
        return None


class GradeAppealListView(APIView):
    """
    GET /api/grades/appeals/
    Permission: role-based filtering
    - STUDENT → own appeals only
    - TEACHER → appeals assigned to them
    - ADMIN/SECTION_ADMIN → all school appeals
    """

    permission_classes = [permissions.IsAuthenticated]

    @extend_schema(tags=["grades"], summary="List grade appeals")
    def get(self, request):
        user = request.user
        qs = GradeAppeal.objects.select_related(
            "student__user",
            "grade__exam_type__subject",
            "subject_average__subject",
            "trimester_average",
            "assigned_to_teacher__user",
            "responded_by",
        ).order_by("-created_at")

        if user.role == User.Role.STUDENT:
            try:
                student = user.student_profile
                qs = qs.filter(student=student)
            except StudentProfile.DoesNotExist:
                qs = qs.none()

        elif user.role == User.Role.TEACHER:
            try:
                teacher_profile = user.teacher_profile
                qs = qs.filter(assigned_to_teacher=teacher_profile)
            except Exception:
                qs = qs.none()

        elif user.role in (
            User.Role.ADMIN,
            User.Role.SECTION_ADMIN,
            User.Role.SUPER_ADMIN,
        ):
            school = _resolve_school(request)
            if school:
                qs = qs.filter(student__current_class__section__school=school)
        else:
            qs = qs.none()

        # Optional filters
        appeal_status = request.query_params.get("status")
        if appeal_status:
            qs = qs.filter(status=appeal_status.upper())

        serializer = GradeAppealSerializer(qs, many=True)
        return Response(serializer.data)


class GradeAppealRespondView(APIView):
    """
    POST /api/grades/appeals/{id}/respond/
    Body: { status, response, corrected_value (optional) }
    Permission: TEACHER (for their appeals) or ADMIN
    """

    permission_classes = [permissions.IsAuthenticated]

    @extend_schema(
        tags=["grades"],
        summary="Respond to a grade appeal",
        request=GradeAppealRespondSerializer,
    )
    def post(self, request, pk):
        ser = GradeAppealRespondSerializer(data=request.data)
        ser.is_valid(raise_exception=True)
        data = ser.validated_data

        appeal = get_object_or_404(
            GradeAppeal.objects.select_related(
                "grade__exam_type",
                "subject_average",
                "trimester_average",
                "student__user",
            ),
            pk=pk,
        )

        user = request.user

        # Permission check: teacher can only respond to their own assigned appeals
        if user.role == User.Role.TEACHER:
            try:
                tp = user.teacher_profile
                if appeal.assigned_to_teacher_id != tp.pk:
                    return Response(
                        {"detail": "Ce recours ne vous est pas assigné."},
                        status=status.HTTP_403_FORBIDDEN,
                    )
            except Exception:
                return Response(
                    {"detail": "Profil enseignant introuvable."},
                    status=status.HTTP_403_FORBIDDEN,
                )
        elif user.role not in (
            User.Role.ADMIN,
            User.Role.SECTION_ADMIN,
            User.Role.SUPER_ADMIN,
        ):
            return Response(
                {"detail": "Permission refusée."},
                status=status.HTTP_403_FORBIDDEN,
            )

        # Update appeal
        appeal.status = data["status"]
        appeal.response = data["response"]
        appeal.responded_by = user
        appeal.responded_at = timezone.now()

        # If ACCEPTED + corrected_value → apply correction
        corrected_value = data.get("corrected_value")
        if (
            data["status"] == GradeAppeal.Status.ACCEPTED
            and corrected_value is not None
        ):
            appeal.corrected_value = corrected_value

            if appeal.appeal_type == GradeAppeal.AppealType.EXAM_GRADE and appeal.grade:
                # Correct the grade directly
                appeal.grade.score = corrected_value
                appeal.grade.is_absent = False
                appeal.grade.entered_by = user
                appeal.grade.save()  # triggers cascade via signal

            elif (
                appeal.appeal_type == GradeAppeal.AppealType.SUBJECT_AVERAGE
                and appeal.subject_average
            ):
                from .services import LockedException, admin_override_average

                try:
                    admin_override_average(
                        override_type="subject_average",
                        object_id=appeal.subject_average_id,
                        new_value=corrected_value,
                        admin_user=user,
                    )
                except (LockedException, Exception) as e:
                    logger.warning("Appeal override failed: %s", e)

            elif (
                appeal.appeal_type == GradeAppeal.AppealType.TRIMESTER_AVERAGE
                and appeal.trimester_average
            ):
                from .services import LockedException, admin_override_average

                try:
                    admin_override_average(
                        override_type="trimester_average",
                        object_id=appeal.trimester_average_id,
                        new_value=corrected_value,
                        admin_user=user,
                    )
                except (LockedException, Exception) as e:
                    logger.warning("Appeal override failed: %s", e)

        appeal.save()

        # Audit
        audit_action = (
            "APPEAL_ACCEPTED"
            if data["status"] == GradeAppeal.Status.ACCEPTED
            else "APPEAL_REJECTED"
        )
        log_grade_action(
            action=audit_action,
            instance=appeal,
            performed_by=user,
            request=request,
            student=appeal.student,
            old_value=appeal.original_value,
            new_value=corrected_value,
            reason=data["response"],
        )

        # Notify student of outcome
        try:
            from apps.notifications.tasks import send_notification

            status_label = (
                "accepté" if data["status"] == GradeAppeal.Status.ACCEPTED else "rejeté"
            )
            send_notification.delay(
                user_id=str(appeal.student.user_id),
                title=f"Recours {status_label}",
                body=f"Votre recours a été {status_label}. Réponse : {data['response'][:100]}",
                notification_type="GRADE_APPEAL",
                related_object_id=str(appeal.pk),
                related_object_type="GradeAppeal",
            )
        except Exception:
            logger.exception("Failed to notify student about appeal response")

        return Response(GradeAppealSerializer(appeal).data)


class GradeAppealPendingCountView(APIView):
    """
    GET /api/grades/appeals/pending-count/
    Returns { teacher_pending, admin_pending }
    For sidebar badge notifications.
    """

    permission_classes = [permissions.IsAuthenticated]

    @extend_schema(tags=["grades"], summary="Pending appeal counts")
    def get(self, request):
        user = request.user
        school = _resolve_school(request)

        base_qs = GradeAppeal.objects.filter(
            status__in=[GradeAppeal.Status.PENDING, GradeAppeal.Status.UNDER_REVIEW],
        )
        if school:
            base_qs = base_qs.filter(student__current_class__section__school=school)

        teacher_pending = 0
        admin_pending = 0

        if user.role == User.Role.TEACHER:
            try:
                tp = user.teacher_profile
                teacher_pending = base_qs.filter(assigned_to_teacher=tp).count()
            except Exception:
                pass
        elif user.role in (
            User.Role.ADMIN,
            User.Role.SECTION_ADMIN,
            User.Role.SUPER_ADMIN,
        ):
            teacher_pending = base_qs.filter(
                assigned_to_teacher__isnull=False,
            ).count()
            admin_pending = base_qs.filter(assigned_to_admin=True).count()

        return Response(
            {
                "teacher_pending": teacher_pending,
                "admin_pending": admin_pending,
            }
        )


# ═══════════════════════════════════════════════════════════════════════════
# 6. AUDIT LOG
# ═══════════════════════════════════════════════════════════════════════════


class GradeAuditLogView(APIView):
    """
    GET /api/grades/audit-log/?student_id=X&trimester=1
    Permission: IsAdminOnly ou enseignant de la classe.
    Retourne l'historique complet des modifications d'un élève.
    """

    permission_classes = [permissions.IsAuthenticated]

    @extend_schema(
        tags=["grades"],
        summary="Audit log timeline for a student",
    )
    def get(self, request):
        student_id = request.query_params.get("student_id")
        if not student_id:
            return Response(
                {"detail": "Le paramètre student_id est requis."},
                status=status.HTTP_400_BAD_REQUEST,
            )

        qs = GradeAuditLog.objects.filter(
            student_id=student_id,
        ).select_related("performed_by")

        # Filters
        trimester = request.query_params.get("trimester")
        action_filter = request.query_params.get("action")
        if trimester:
            qs = qs.filter(trimester=int(trimester))
        if action_filter:
            qs = qs.filter(action=action_filter.upper())

        # Tenant isolation
        school = _resolve_school(request)
        if school:
            qs = qs.filter(student__current_class__section__school=school)

        # Permission: teacher can only see logs for students in their classes
        user = request.user
        if user.role == User.Role.TEACHER:
            teacher_class_ids = TeacherAssignment.objects.filter(
                teacher=user,
            ).values_list("assigned_class_id", flat=True)
            qs = qs.filter(student__current_class_id__in=teacher_class_ids)
        elif user.role not in (
            User.Role.ADMIN,
            User.Role.SECTION_ADMIN,
            User.Role.SUPER_ADMIN,
        ):
            return Response(
                {"detail": "Permission refusée."},
                status=status.HTTP_403_FORBIDDEN,
            )

        serializer = GradeAuditLogSerializer(qs[:200], many=True)
        return Response(serializer.data)


# ═══════════════════════════════════════════════════════════════════════════
# 7. GRADE WORKFLOW — submit / return
# ═══════════════════════════════════════════════════════════════════════════


class GradeSubmitView(APIView):
    """
    POST /api/grades/submit/
    Body: { exam_type_id }
    Permission: IsTeacherOfClassroom

    Workflow: DRAFT/RETURNED → SUBMITTED (for admin review)
    """

    permission_classes = [
        permissions.IsAuthenticated,
        IsTeacherOfClassroom,
    ]

    @extend_schema(
        tags=["grades"],
        summary="Submit grades for admin review",
        request=GradeSubmitSerializer,
    )
    def post(self, request):
        ser = GradeSubmitSerializer(data=request.data)
        ser.is_valid(raise_exception=True)

        from .services import InvalidTransitionError, submit_grades

        try:
            result = submit_grades(
                exam_type_id=ser.validated_data["exam_type_id"],
                teacher_user=request.user,
            )
        except InvalidTransitionError as e:
            return Response(
                {"detail": str(e)},
                status=status.HTTP_400_BAD_REQUEST,
            )

        # Audit
        exam_type = get_object_or_404(ExamType, pk=ser.validated_data["exam_type_id"])
        log_grade_action(
            action="GRADE_SUBMITTED",
            instance=exam_type,
            performed_by=request.user,
            request=request,
            subject_name=str(exam_type.subject),
            exam_name=exam_type.name,
            trimester=exam_type.trimester,
        )

        # Notify admins
        try:
            from apps.notifications.tasks import send_notification

            school = exam_type.classroom.section.school
            admins = User.objects.filter(
                school=school,
                role__in=[User.Role.ADMIN, User.Role.SECTION_ADMIN],
                is_active=True,
            )
            for admin_user in admins:
                send_notification.delay(
                    user_id=str(admin_user.pk),
                    title="Notes soumises pour validation",
                    body=(
                        f"{request.user.get_full_name()} a soumis les notes de "
                        f"{exam_type.subject.name} ({exam_type.name}) "
                        f"pour la classe {exam_type.classroom.name}."
                    ),
                    notification_type="GRADE_SUBMITTED",
                    related_object_id=str(exam_type.pk),
                    related_object_type="ExamType",
                )
        except Exception:
            logger.exception("Failed to send submission notifications")

        return Response(result)


class GradeReturnView(APIView):
    """
    POST /api/grades/return/
    Body: { exam_type_id, comment }
    Permission: IsAdminOnly

    Workflow: SUBMITTED → RETURNED (with admin comment)
    """

    permission_classes = [permissions.IsAuthenticated, IsAdminOnly]

    @extend_schema(
        tags=["grades"],
        summary="Return grades to teacher with comment",
        request=GradeReturnSerializer,
    )
    def post(self, request):
        ser = GradeReturnSerializer(data=request.data)
        ser.is_valid(raise_exception=True)

        from .services import InvalidTransitionError, return_grades

        try:
            result = return_grades(
                exam_type_id=ser.validated_data["exam_type_id"],
                admin_user=request.user,
                comment=ser.validated_data["comment"],
            )
        except (InvalidTransitionError, ValueError) as e:
            return Response(
                {"detail": str(e)},
                status=status.HTTP_400_BAD_REQUEST,
            )

        # Audit
        exam_type = get_object_or_404(ExamType, pk=ser.validated_data["exam_type_id"])
        log_grade_action(
            action="GRADE_RETURNED",
            instance=exam_type,
            performed_by=request.user,
            request=request,
            reason=ser.validated_data["comment"],
            subject_name=str(exam_type.subject),
            exam_name=exam_type.name,
            trimester=exam_type.trimester,
        )

        return Response(result)


class GradeWorkflowStatusView(APIView):
    """
    GET /api/grades/workflow-status/?exam_type_id=X
    Returns grade counts per status for an exam type.
    """

    permission_classes = [permissions.IsAuthenticated, IsAdminOrTeacherOfClassroom]

    @extend_schema(tags=["grades"], summary="Grade workflow status counts")
    def get(self, request):
        exam_type_id = request.query_params.get("exam_type_id")
        if not exam_type_id:
            return Response(
                {"detail": "exam_type_id requis."},
                status=status.HTTP_400_BAD_REQUEST,
            )

        exam_type = get_object_or_404(ExamType, pk=exam_type_id)
        grades = Grade.objects.filter(exam_type=exam_type)

        counts = {
            "total": grades.count(),
            "draft": grades.filter(status=Grade.Status.DRAFT).count(),
            "submitted": grades.filter(status=Grade.Status.SUBMITTED).count(),
            "published": grades.filter(status=Grade.Status.PUBLISHED).count(),
            "returned": grades.filter(status=Grade.Status.RETURNED).count(),
        }

        return Response(counts)


# ═══════════════════════════════════════════════════════════════════════════
# 8. STUDENT AVERAGES ENDPOINT
# ═══════════════════════════════════════════════════════════════════════════


class StudentAveragesView(APIView):
    """
    GET /api/grades/students/{student_id}/averages/?academic_year_id=X

    Returns a comprehensive view of a student's averages:
    - Subject averages per trimester
    - Trimester averages with rankings
    - Annual averages
    """

    permission_classes = [permissions.IsAuthenticated]

    @extend_schema(tags=["grades"], summary="Get student averages overview")
    def get(self, request, student_id):
        from .services import get_student_averages

        academic_year_id = request.query_params.get("academic_year_id")

        # Permission: student sees own, teacher sees their class, admin sees school
        user = request.user
        if user.role == User.Role.STUDENT:
            try:
                if str(user.student_profile.pk) != str(student_id):
                    return Response(
                        {"detail": "Accès interdit."},
                        status=status.HTTP_403_FORBIDDEN,
                    )
            except Exception:
                return Response(
                    {"detail": "Profil étudiant introuvable."},
                    status=status.HTTP_403_FORBIDDEN,
                )
        elif user.role == User.Role.PARENT:
            # Parent can only see their children
            from apps.academics.models import StudentProfile

            child_ids = StudentProfile.objects.filter(
                parent_user=user,
                is_deleted=False,
            ).values_list("pk", flat=True)
            if student_id not in [str(cid) for cid in child_ids]:
                return Response(
                    {"detail": "Accès interdit."},
                    status=status.HTTP_403_FORBIDDEN,
                )
        elif user.role == User.Role.TEACHER:
            # Teacher can see students in their classes
            teacher_class_ids = TeacherAssignment.objects.filter(
                teacher=user,
            ).values_list("assigned_class_id", flat=True)
            from apps.academics.models import StudentProfile

            if not StudentProfile.objects.filter(
                pk=student_id,
                current_class_id__in=teacher_class_ids,
            ).exists():
                return Response(
                    {"detail": "Accès interdit."},
                    status=status.HTTP_403_FORBIDDEN,
                )
        elif user.role not in (
            User.Role.ADMIN,
            User.Role.SECTION_ADMIN,
            User.Role.SUPER_ADMIN,
        ):
            return Response(
                {"detail": "Permission refusée."},
                status=status.HTTP_403_FORBIDDEN,
            )

        result = get_student_averages(student_id, academic_year_id)
        return Response(result)


# ═══════════════════════════════════════════════════════════════════════════
# 9. CSV GRADE IMPORT
# ═══════════════════════════════════════════════════════════════════════════


class CSVImportPreviewView(APIView):
    """
    POST /api/grades/csv-import/preview/
    Body: multipart/form-data with exam_type_id + csv_file
    Permission: IsTeacherOfClassroom

    Parses CSV and returns a preview of matched/unmatched students.
    Does NOT save anything.
    """

    permission_classes = [
        permissions.IsAuthenticated,
        IsTeacherOfClassroom,
    ]

    @extend_schema(
        tags=["grades"],
        summary="Preview CSV grade import",
        request=CSVImportPreviewSerializer,
    )
    def post(self, request):
        ser = CSVImportPreviewSerializer(data=request.data)
        ser.is_valid(raise_exception=True)

        csv_file = ser.validated_data["csv_file"]

        # Read file content
        try:
            content = csv_file.read().decode("utf-8-sig")  # Handle BOM
        except UnicodeDecodeError:
            try:
                csv_file.seek(0)
                content = csv_file.read().decode("latin-1")
            except Exception:
                return Response(
                    {"detail": "Impossible de lire le fichier. Encodage non supporté."},
                    status=status.HTTP_400_BAD_REQUEST,
                )

        from .services import CSVImportError, parse_csv_grades

        try:
            result = parse_csv_grades(
                csv_content=content,
                exam_type_id=ser.validated_data["exam_type_id"],
                teacher_user=request.user,
            )
        except CSVImportError as e:
            return Response(
                {"detail": str(e)},
                status=status.HTTP_400_BAD_REQUEST,
            )

        return Response(result)


class CSVImportConfirmView(APIView):
    """
    POST /api/grades/csv-import/confirm/
    Body: { exam_type_id, matched: [...] }
    Permission: IsTeacherOfClassroom

    Confirms a preview and saves grades as DRAFT.
    Teacher can then modify manually before submitting.
    """

    permission_classes = [
        permissions.IsAuthenticated,
        IsTeacherOfClassroom,
    ]

    @extend_schema(
        tags=["grades"],
        summary="Confirm CSV grade import",
        request=CSVImportConfirmSerializer,
    )
    def post(self, request):
        ser = CSVImportConfirmSerializer(data=request.data)
        ser.is_valid(raise_exception=True)

        from .services import confirm_csv_import

        result = confirm_csv_import(
            preview_data=ser.validated_data,
            teacher_user=request.user,
        )

        # Audit
        exam_type = get_object_or_404(ExamType, pk=ser.validated_data["exam_type_id"])
        log_grade_action(
            action="GRADE_CSV_IMPORTED",
            instance=exam_type,
            performed_by=request.user,
            request=request,
            subject_name=str(exam_type.subject),
            exam_name=exam_type.name,
            trimester=exam_type.trimester,
        )

        return Response(result)


# ═══════════════════════════════════════════════════════════════════════════
# 10. TRIMESTER CONFIG — CRUD for per-school configuration
# ═══════════════════════════════════════════════════════════════════════════


class TrimesterConfigView(APIView):
    """
    GET / PUT /api/grades/trimester-config/
    Permission: IsAdminOnly

    Get or update the trimester configuration for the current school.
    Creates a default config if none exists.
    """

    permission_classes = [permissions.IsAuthenticated, IsAdminOnly]

    @extend_schema(tags=["grades"], summary="Get trimester configuration")
    def get(self, request):
        school = _resolve_school(request)
        if not school:
            return Response(
                {"detail": "École non trouvée."},
                status=status.HTTP_404_NOT_FOUND,
            )

        config, _ = TrimesterConfig.objects.get_or_create(school=school)
        return Response(TrimesterConfigSerializer(config).data)

    @extend_schema(
        tags=["grades"],
        summary="Update trimester configuration",
        request=TrimesterConfigSerializer,
    )
    def put(self, request):
        school = _resolve_school(request)
        if not school:
            return Response(
                {"detail": "École non trouvée."},
                status=status.HTTP_404_NOT_FOUND,
            )

        config, _ = TrimesterConfig.objects.get_or_create(school=school)
        ser = TrimesterConfigSerializer(config, data=request.data, partial=True)
        ser.is_valid(raise_exception=True)
        ser.save()
        return Response(ser.data)


# ═══════════════════════════════════════════════════════════════════════════
# 11. REPORT CARD TEMPLATES — CRUD
# ═══════════════════════════════════════════════════════════════════════════


class ReportCardTemplateListCreateView(APIView):
    """
    GET  /api/grades/report-card-templates/
    POST /api/grades/report-card-templates/
    """

    permission_classes = [permissions.IsAuthenticated, IsAdminOnly]

    @extend_schema(tags=["grades"], summary="List report card templates")
    def get(self, request):
        school = _resolve_school(request)
        if not school:
            return Response({"detail": "École non trouvée."}, status=404)

        templates = ReportCardTemplate.objects.filter(school=school)
        return Response(ReportCardTemplateSerializer(templates, many=True).data)

    @extend_schema(
        tags=["grades"],
        summary="Create report card template",
        request=ReportCardTemplateSerializer,
    )
    def post(self, request):
        school = _resolve_school(request)
        if not school:
            return Response({"detail": "École non trouvée."}, status=404)

        ser = ReportCardTemplateSerializer(data={**request.data, "school": str(school.pk)})
        ser.is_valid(raise_exception=True)
        ser.save()
        return Response(ser.data, status=status.HTTP_201_CREATED)


class ReportCardTemplateDetailView(APIView):
    """
    GET/PUT/DELETE /api/grades/report-card-templates/{id}/
    """

    permission_classes = [permissions.IsAuthenticated, IsAdminOnly]

    def _get_template(self, pk, request):
        school = _resolve_school(request)
        return get_object_or_404(ReportCardTemplate, pk=pk, school=school)

    @extend_schema(tags=["grades"], summary="Get report card template")
    def get(self, request, pk):
        template = self._get_template(pk, request)
        return Response(ReportCardTemplateSerializer(template).data)

    @extend_schema(tags=["grades"], summary="Update report card template")
    def put(self, request, pk):
        template = self._get_template(pk, request)
        ser = ReportCardTemplateSerializer(template, data=request.data, partial=True)
        ser.is_valid(raise_exception=True)
        ser.save()
        return Response(ser.data)

    @extend_schema(tags=["grades"], summary="Delete report card template")
    def delete(self, request, pk):
        template = self._get_template(pk, request)
        template.delete()
        return Response(status=status.HTTP_204_NO_CONTENT)


# ═══════════════════════════════════════════════════════════════════════════
# 12. REPORT CARD GENERATION (class & school-wide with progress)
# ═══════════════════════════════════════════════════════════════════════════


class GenerateClassReportCardsView(APIView):
    """POST /api/grades/report-cards/generate-class/"""

    permission_classes = [permissions.IsAuthenticated, IsAdminOnly]

    @extend_schema(
        tags=["grades"],
        summary="Generate report cards for a class",
        request=GenerateClassReportCardsSerializer,
    )
    def post(self, request):
        ser = GenerateClassReportCardsSerializer(data=request.data)
        ser.is_valid(raise_exception=True)

        from .tasks import generate_class_report_cards

        task = generate_class_report_cards.delay(
            class_id=str(ser.validated_data["class_id"]),
            trimester=ser.validated_data["trimester"],
            academic_year_id=str(ser.validated_data["academic_year_id"]),
        )

        return Response({
            "task_id": task.id,
            "status": "started",
            "message": "Génération des bulletins lancée.",
        })


class GenerateSchoolReportCardsView(APIView):
    """POST /api/grades/report-cards/generate-school/"""

    permission_classes = [permissions.IsAuthenticated, IsAdminOnly]

    @extend_schema(
        tags=["grades"],
        summary="Generate report cards for entire school",
        request=GenerateSchoolReportCardsSerializer,
    )
    def post(self, request):
        school = _resolve_school(request)
        if not school:
            return Response({"detail": "École non trouvée."}, status=404)

        ser = GenerateSchoolReportCardsSerializer(data=request.data)
        ser.is_valid(raise_exception=True)

        from .tasks import generate_school_report_cards

        task = generate_school_report_cards.delay(
            school_id=str(school.pk),
            academic_year_id=str(ser.validated_data["academic_year_id"]),
            trimester=ser.validated_data["trimester"],
            template_id=str(ser.validated_data.get("template_id")) if ser.validated_data.get("template_id") else None,
            send_to_parents=ser.validated_data.get("send_to_parents", False),
        )

        return Response({
            "task_id": task.id,
            "status": "started",
            "message": "Génération des bulletins de l'école lancée.",
        })


class ReportCardProgressView(APIView):
    """GET /api/grades/report-cards/progress/{task_id}/"""

    permission_classes = [permissions.IsAuthenticated, IsAdminOnly]

    @extend_schema(tags=["grades"], summary="Check report card generation progress")
    def get(self, request, task_id):
        from django.core.cache import cache

        progress_key = f"report_cards_progress_{task_id}"
        progress = cache.get(progress_key)

        if progress is None:
            return Response({
                "status": "unknown",
                "message": "Tâche non trouvée ou expirée.",
            })

        return Response(progress)


# ═══════════════════════════════════════════════════════════════════════════
# 13. GRADE ANALYTICS — comprehensive analytics API
# ═══════════════════════════════════════════════════════════════════════════


class GradeAnalyticsView(APIView):
    """
    GET /api/grades/analytics/
    Query params: academic_year_id, trimester (optional)

    Returns comprehensive analytics:
    - Subject averages across classes
    - Pass/fail rates by class, level, section
    - Class performance comparison
    - Top 10 students by level
    - Trimester average evolution (T1→T2→T3)
    - At-risk students (negative trajectory)
    - Teacher analytics (class averages, grade submission rates)
    """

    permission_classes = [permissions.IsAuthenticated, IsAdminOnly]

    @extend_schema(tags=["grades"], summary="Comprehensive grade analytics")
    def get(self, request):
        from decimal import Decimal as D

        from django.db.models import Avg, Count, F, Q

        school = _resolve_school(request)
        if not school:
            return Response({"detail": "École non trouvée."}, status=404)

        academic_year_id = request.query_params.get("academic_year_id")
        trimester = request.query_params.get("trimester")

        # Determine academic year
        if academic_year_id:
            from apps.schools.models import AcademicYear
            academic_year = AcademicYear.objects.filter(pk=academic_year_id).first()
        else:
            academic_year = _get_active_academic_year(school)

        if not academic_year:
            return Response({"detail": "Année académique non trouvée."}, status=404)

        from apps.academics.models import Class

        classes = Class.objects.filter(
            section__school=school,
            academic_year=academic_year,
        ).select_related("level", "section")

        # Base query filters
        base_filter = {"classroom__in": classes, "academic_year": academic_year}
        if trimester:
            base_filter["trimester"] = int(trimester)

        # 1. Subject averages across classes
        subject_averages = (
            SubjectAverage.objects.filter(**base_filter)
            .values("subject__name", "classroom__name")
            .annotate(avg=Avg("calculated_average"))
            .order_by("subject__name", "classroom__name")
        )

        subject_data = {}
        for sa in subject_averages:
            subj = sa["subject__name"]
            if subj not in subject_data:
                subject_data[subj] = []
            subject_data[subj].append({
                "class": sa["classroom__name"],
                "average": round(float(sa["avg"] or 0), 2),
            })

        # 2. Pass/fail rates
        trim_filter = {"classroom__in": classes, "academic_year": academic_year}
        if trimester:
            trim_filter["trimester"] = int(trimester)

        pass_threshold = D("10")
        try:
            tc = TrimesterConfig.objects.get(school=school)
            pass_threshold = tc.pass_threshold
        except TrimesterConfig.DoesNotExist:
            pass

        pass_fail_by_class = (
            TrimesterAverage.objects.filter(**trim_filter)
            .values("classroom__name")
            .annotate(
                total=Count("id"),
                passed=Count("id", filter=Q(calculated_average__gte=pass_threshold)),
                failed=Count("id", filter=Q(calculated_average__lt=pass_threshold)),
            )
            .order_by("classroom__name")
        )

        pass_fail_data = []
        for pf in pass_fail_by_class:
            total = pf["total"] or 1
            pass_fail_data.append({
                "class": pf["classroom__name"],
                "total": pf["total"],
                "passed": pf["passed"],
                "failed": pf["failed"],
                "pass_rate": round(pf["passed"] / total * 100, 1),
            })

        # 3. Class performance comparison
        class_comparison = (
            TrimesterAverage.objects.filter(**trim_filter)
            .values("classroom__name")
            .annotate(
                avg=Avg("calculated_average"),
                count=Count("id"),
            )
            .order_by("-avg")
        )

        class_comparison_data = [
            {
                "class": cc["classroom__name"],
                "average": round(float(cc["avg"] or 0), 2),
                "students": cc["count"],
            }
            for cc in class_comparison
        ]

        # 4. Top 10 students by level
        top_students = []
        levels = classes.values_list("level", flat=True).distinct()
        for level_id in levels:
            level_trim = TrimesterAverage.objects.filter(
                classroom__level_id=level_id,
                academic_year=academic_year,
            )
            if trimester:
                level_trim = level_trim.filter(trimester=int(trimester))

            top = (
                level_trim.order_by("-calculated_average")
                .select_related("student__user", "classroom__level")[:10]
            )
            level_name = top[0].classroom.level.name if top else ""
            top_students.append({
                "level": level_name,
                "students": [
                    {
                        "name": t.student.user.get_full_name(),
                        "class": t.classroom.name,
                        "average": float(t.effective_average),
                        "rank": t.rank_in_level,
                    }
                    for t in top
                ],
            })

        # 5. Trimester evolution (T1→T2→T3 curves)
        evolution = (
            TrimesterAverage.objects.filter(
                classroom__in=classes,
                academic_year=academic_year,
            )
            .values("trimester")
            .annotate(avg=Avg("calculated_average"), count=Count("id"))
            .order_by("trimester")
        )
        evolution_data = [
            {
                "trimester": e["trimester"],
                "average": round(float(e["avg"] or 0), 2),
                "students": e["count"],
            }
            for e in evolution
        ]

        # 6. At-risk students (negative trajectory T1→T2 or T2→T3)
        at_risk = []
        trim_nums = [1, 2, 3] if not trimester else [int(trimester)]
        if len(trim_nums) > 1 or not trimester:
            from apps.academics.models import StudentProfile

            student_ids = (
                TrimesterAverage.objects.filter(
                    classroom__in=classes, academic_year=academic_year
                )
                .values_list("student_id", flat=True)
                .distinct()
            )

            for sid in student_ids[:500]:  # Limit for performance
                trims = (
                    TrimesterAverage.objects.filter(
                        student_id=sid,
                        academic_year=academic_year,
                        classroom__in=classes,
                    )
                    .order_by("trimester")
                    .values_list("trimester", "calculated_average")
                )
                trim_dict = {t: float(a) for t, a in trims if a is not None}

                if len(trim_dict) >= 2:
                    sorted_trims = sorted(trim_dict.keys())
                    last = sorted_trims[-1]
                    prev = sorted_trims[-2]
                    drop = trim_dict[prev] - trim_dict[last]

                    if drop >= 2.0:  # Significant drop
                        try:
                            student = StudentProfile.objects.select_related(
                                "user", "current_class"
                            ).get(pk=sid)
                            at_risk.append({
                                "student_id": str(sid),
                                "name": student.user.get_full_name(),
                                "class": student.current_class.name if student.current_class else "",
                                "previous_avg": trim_dict[prev],
                                "current_avg": trim_dict[last],
                                "drop": round(drop, 2),
                                "trimesters": f"T{prev}→T{last}",
                            })
                        except StudentProfile.DoesNotExist:
                            pass

            at_risk.sort(key=lambda x: x["drop"], reverse=True)
            at_risk = at_risk[:50]  # Top 50 at-risk

        # 7. Teacher analytics
        teacher_analytics = []
        from apps.academics.models import TeacherAssignment as TA

        assignments = TA.objects.filter(
            assigned_class__in=classes,
        ).select_related("teacher", "subject")

        teacher_map = {}
        for a in assignments:
            tid = str(a.teacher_id)
            if tid not in teacher_map:
                teacher_map[tid] = {
                    "name": a.teacher.get_full_name(),
                    "subjects": [],
                    "classes": [],
                }
            teacher_map[tid]["subjects"].append(a.subject.name if a.subject else "")
            teacher_map[tid]["classes"].append(a.assigned_class.name)

        for tid, info in teacher_map.items():
            # Get class averages for this teacher's classes
            t_classes = info["classes"]
            class_avgs = (
                TrimesterAverage.objects.filter(
                    classroom__name__in=t_classes,
                    academic_year=academic_year,
                )
                .values("classroom__name")
                .annotate(avg=Avg("calculated_average"))
            )

            # Grade submission rate
            t_subjects = info["subjects"]
            total_grades = Grade.objects.filter(
                exam_type__classroom__name__in=t_classes,
                exam_type__subject__name__in=t_subjects,
                exam_type__academic_year=academic_year,
            ).count()
            published = Grade.objects.filter(
                exam_type__classroom__name__in=t_classes,
                exam_type__subject__name__in=t_subjects,
                exam_type__academic_year=academic_year,
                status=Grade.Status.PUBLISHED,
            ).count()

            teacher_analytics.append({
                "teacher_id": tid,
                "name": info["name"],
                "subjects": list(set(info["subjects"])),
                "classes": list(set(info["classes"])),
                "class_averages": {
                    ca["classroom__name"]: round(float(ca["avg"] or 0), 2)
                    for ca in class_avgs
                },
                "total_grades": total_grades,
                "published_grades": published,
                "submission_rate": round(published / total_grades * 100, 1) if total_grades else 0,
            })

        return Response({
            "subject_averages": subject_data,
            "pass_fail_rates": pass_fail_data,
            "class_comparison": class_comparison_data,
            "top_students_by_level": top_students,
            "trimester_evolution": evolution_data,
            "at_risk_students": at_risk,
            "teacher_analytics": teacher_analytics,
        })


# ═══════════════════════════════════════════════════════════════════════════
# 14. MEN OFFICIAL EXPORT FORMAT
# ═══════════════════════════════════════════════════════════════════════════


class MENExportView(APIView):
    """
    GET /api/grades/men-export/
    Query params: academic_year_id, trimester, class_id (optional)

    Exports grade data in official MEN (Ministère de l'Éducation Nationale) format.
    Returns Excel file for BEP/BEM/BAC exam preparation.
    """

    permission_classes = [permissions.IsAuthenticated, IsAdminOnly]

    @extend_schema(tags=["grades"], summary="Export grades in MEN official format")
    def get(self, request):
        import openpyxl
        from django.http import HttpResponse
        from openpyxl.styles import Alignment, Border, Font, PatternFill, Side

        school = _resolve_school(request)
        if not school:
            return Response({"detail": "École non trouvée."}, status=404)

        academic_year_id = request.query_params.get("academic_year_id")
        trimester = request.query_params.get("trimester")
        class_id = request.query_params.get("class_id")

        if not academic_year_id or not trimester:
            return Response(
                {"detail": "academic_year_id et trimester sont requis."},
                status=400,
            )

        from apps.academics.models import Class, LevelSubject, StudentProfile
        from apps.schools.models import AcademicYear

        academic_year = get_object_or_404(AcademicYear, pk=academic_year_id)

        if class_id:
            classes = Class.objects.filter(pk=class_id, section__school=school)
        else:
            classes = Class.objects.filter(
                section__school=school,
                academic_year=academic_year,
            )

        classes = classes.select_related("level", "section")

        wb = openpyxl.Workbook()
        wb.remove(wb.active)

        header_font = Font(bold=True, size=12)
        header_fill = PatternFill(start_color="1F4E79", end_color="1F4E79", fill_type="solid")
        header_font_white = Font(bold=True, color="FFFFFF", size=10)
        thin_border = Border(
            left=Side(style="thin"),
            right=Side(style="thin"),
            top=Side(style="thin"),
            bottom=Side(style="thin"),
        )

        for class_obj in classes:
            ws = wb.create_sheet(title=class_obj.name[:31])

            # Header
            ws.merge_cells("A1:H1")
            ws["A1"] = "الجمهورية الجزائرية الديمقراطية الشعبية"
            ws["A1"].font = header_font
            ws["A1"].alignment = Alignment(horizontal="center")

            ws.merge_cells("A2:H2")
            ws["A2"] = f"وزارة التربية الوطنية — {school.name}"
            ws["A2"].alignment = Alignment(horizontal="center")

            ws.merge_cells("A3:H3")
            ws["A3"] = f"كشف النقاط — القسم: {class_obj.name} — الفصل {trimester} — {academic_year.name}"
            ws["A3"].font = Font(bold=True, size=11)
            ws["A3"].alignment = Alignment(horizontal="center")

            # Get subjects for this level
            subjects = LevelSubject.objects.filter(
                level=class_obj.level,
                stream=class_obj.stream,
            ).select_related("subject").order_by("subject__name")

            # Column headers
            headers = ["الرقم", "الاسم واللقب"]
            for ls in subjects:
                headers.append(f"{ls.subject.name}\n(/{ls.coefficient})")
            headers += ["المعدل", "الترتيب", "التقدير"]

            for col, h in enumerate(headers, 1):
                cell = ws.cell(row=5, column=col, value=h)
                cell.font = header_font_white
                cell.fill = header_fill
                cell.border = thin_border
                cell.alignment = Alignment(horizontal="center", wrap_text=True)

            # Students data
            students = (
                StudentProfile.objects.filter(
                    current_class=class_obj,
                    is_deleted=False,
                )
                .select_related("user")
                .order_by("user__last_name_ar", "user__first_name_ar")
            )

            for row_idx, student in enumerate(students, 6):
                ws.cell(row=row_idx, column=1, value=row_idx - 5).border = thin_border
                full_name = (
                    f"{getattr(student.user, 'last_name_ar', student.user.last_name)} "
                    f"{getattr(student.user, 'first_name_ar', student.user.first_name)}"
                )
                ws.cell(row=row_idx, column=2, value=full_name).border = thin_border

                for col_idx, ls in enumerate(subjects, 3):
                    try:
                        sa = SubjectAverage.objects.get(
                            student=student,
                            subject=ls.subject,
                            classroom=class_obj,
                            academic_year=academic_year,
                            trimester=int(trimester),
                        )
                        val = float(sa.effective_average)
                    except SubjectAverage.DoesNotExist:
                        val = ""
                    cell = ws.cell(row=row_idx, column=col_idx, value=val)
                    cell.border = thin_border
                    cell.alignment = Alignment(horizontal="center")

                # Trimester average, rank, appreciation
                col_avg = len(subjects) + 3
                try:
                    ta = TrimesterAverage.objects.get(
                        student=student,
                        classroom=class_obj,
                        academic_year=academic_year,
                        trimester=int(trimester),
                    )
                    ws.cell(row=row_idx, column=col_avg, value=float(ta.effective_average)).border = thin_border
                    ws.cell(row=row_idx, column=col_avg + 1, value=ta.rank_in_class).border = thin_border
                    ws.cell(row=row_idx, column=col_avg + 2, value=ta.appreciation).border = thin_border
                except TrimesterAverage.DoesNotExist:
                    for c in range(3):
                        ws.cell(row=row_idx, column=col_avg + c, value="").border = thin_border

            # Auto-adjust column widths
            for col in ws.columns:
                max_len = 0
                for cell in col:
                    if cell.value:
                        max_len = max(max_len, len(str(cell.value)))
                ws.column_dimensions[col[0].column_letter].width = max(max_len + 2, 12)

        response = HttpResponse(
            content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        )
        response["Content-Disposition"] = (
            f'attachment; filename="MEN_export_{school.name}_T{trimester}.xlsx"'
        )
        wb.save(response)
        return response
