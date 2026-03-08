"""
Bulk Student Import endpoints.

POST   /api/v1/auth/students/bulk-import/
       Upload an Excel file → kicks off Celery task → returns job ID.

GET    /api/v1/auth/students/bulk-import/<uuid:job_id>/progress/
       Poll job progress (total, processed, created, linked, errors, pct).

GET    /api/v1/auth/students/bulk-import/template/
       Download a pre-built Excel template with the expected columns.
"""

import io
import logging

from django.http import HttpResponse
from drf_spectacular.utils import extend_schema, OpenApiResponse
from rest_framework import permissions, serializers, status
from rest_framework.parsers import MultiPartParser
from rest_framework.response import Response
from rest_framework.views import APIView

from core.permissions import IsSchoolAdmin

from .models import BulkImportJob
from .tasks import bulk_import_students

logger = logging.getLogger(__name__)


# ═══════════════════════════════════════════════════════════════════════════
# Serializers (inline, only used here)
# ═══════════════════════════════════════════════════════════════════════════


class BulkImportUploadSerializer(serializers.Serializer):
    file = serializers.FileField(
        help_text="Excel (.xlsx) file following the import template."
    )


class BulkImportJobSerializer(serializers.ModelSerializer):
    progress_pct = serializers.IntegerField(read_only=True)

    class Meta:
        model = BulkImportJob
        fields = [
            "id",
            "status",
            "total_rows",
            "processed_rows",
            "created_count",
            "linked_count",
            "error_count",
            "error_rows",
            "progress_pct",
            "celery_task_id",
            "created_at",
            "completed_at",
        ]
        read_only_fields = fields


# ═══════════════════════════════════════════════════════════════════════════
# 1. Upload & Start Import
# ═══════════════════════════════════════════════════════════════════════════


@extend_schema(
    tags=["bulk-import"],
    summary="Upload student Excel file and start bulk import",
    request=BulkImportUploadSerializer,
    responses={
        201: BulkImportJobSerializer,
        400: OpenApiResponse(description="Invalid file."),
    },
)
class BulkImportUploadView(APIView):
    """
    POST /api/v1/auth/students/bulk-import/

    Accepts an ``.xlsx`` file, validates the extension, creates a
    ``BulkImportJob`` record, and dispatches the Celery task.
    Returns the job representation so the admin can start polling.
    """

    permission_classes = [permissions.IsAuthenticated, IsSchoolAdmin]
    parser_classes = [MultiPartParser]

    def post(self, request):
        ser = BulkImportUploadSerializer(data=request.data)
        ser.is_valid(raise_exception=True)
        uploaded = ser.validated_data["file"]

        # Basic validation
        if not uploaded.name.endswith(".xlsx"):
            return Response(
                {"detail": "Only .xlsx files are accepted."},
                status=status.HTTP_400_BAD_REQUEST,
            )
        if uploaded.size > 10 * 1024 * 1024:  # 10 MB limit
            return Response(
                {"detail": "File too large (max 10 MB)."},
                status=status.HTTP_400_BAD_REQUEST,
            )

        school = request.user.school
        if not school:
            return Response(
                {"detail": "No school associated with your account."},
                status=status.HTTP_400_BAD_REQUEST,
            )

        # Persist job
        job = BulkImportJob.objects.create(
            school=school,
            uploaded_by=request.user,
            file=uploaded,
        )

        # Dispatch Celery task
        task = bulk_import_students.delay(str(job.pk))
        job.celery_task_id = task.id
        job.save(update_fields=["celery_task_id"])

        logger.info(
            "Bulk import started — job=%s school=%s user=%s",
            job.pk,
            school.pk,
            request.user.pk,
        )

        return Response(
            BulkImportJobSerializer(job).data,
            status=status.HTTP_201_CREATED,
        )


# ═══════════════════════════════════════════════════════════════════════════
# 2. Poll Progress
# ═══════════════════════════════════════════════════════════════════════════


@extend_schema(
    tags=["bulk-import"],
    summary="Check bulk import job progress",
    responses={200: BulkImportJobSerializer},
)
class BulkImportProgressView(APIView):
    """
    GET /api/v1/auth/students/bulk-import/<uuid:job_id>/progress/

    Returns current counters & status so the admin panel can render
    a progress bar.
    """

    permission_classes = [permissions.IsAuthenticated, IsSchoolAdmin]

    def get(self, request, job_id):
        try:
            job = BulkImportJob.objects.get(
                pk=job_id,
                school=request.user.school,
            )
        except BulkImportJob.DoesNotExist:
            return Response(
                {"detail": "Job not found."},
                status=status.HTTP_404_NOT_FOUND,
            )

        return Response(BulkImportJobSerializer(job).data)


# ═══════════════════════════════════════════════════════════════════════════
# 3. Download Excel Template
# ═══════════════════════════════════════════════════════════════════════════


@extend_schema(
    tags=["bulk-import"],
    summary="Download the Excel import template",
    responses={200: OpenApiResponse(description=".xlsx file download")},
)
class BulkImportTemplateView(APIView):
    """
    GET /api/v1/auth/students/bulk-import/template/

    Generates a fresh ``.xlsx`` with the expected header row and
    a sample data row so the admin knows what to fill.
    """

    permission_classes = [permissions.IsAuthenticated, IsSchoolAdmin]

    def get(self, request):
        import openpyxl
        from openpyxl.styles import Font, PatternFill, Alignment

        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "Import Élèves"

        headers = [
            "Prénom élève *",
            "Nom élève *",
            "Date de naissance (YYYY-MM-DD)",
            "Classe *",
            "Cycle * (PRIMARY/MIDDLE/HIGH)",
            "Téléphone parent *",
            "Prénom parent",
            "Nom parent",
            "Lien parent (FATHER/MOTHER/GUARDIAN)",
            "Téléphone 2ème parent",
            "Prénom 2ème parent",
            "Nom 2ème parent",
            "Lien 2ème parent (FATHER/MOTHER/GUARDIAN)",
        ]

        header_font = Font(bold=True, color="FFFFFF", size=11)
        header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")

        for col, header in enumerate(headers, start=1):
            cell = ws.cell(row=1, column=col, value=header)
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = Alignment(horizontal="center", wrap_text=True)
            ws.column_dimensions[cell.column_letter].width = max(len(header) + 4, 18)

        # Sample data row
        sample = [
            "Ahmed", "Benali", "2015-09-15", "3AP-A", "PRIMARY",
            "0555123456", "Karim", "Benali", "FATHER",
            "0555654321", "Fatima", "Benali", "MOTHER",
        ]
        for col, val in enumerate(sample, start=1):
            ws.cell(row=2, column=col, value=val)

        # Write to response
        buf = io.BytesIO()
        wb.save(buf)
        buf.seek(0)

        response = HttpResponse(
            buf.getvalue(),
            content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        )
        response["Content-Disposition"] = (
            'attachment; filename="import_eleves_template.xlsx"'
        )
        return response
